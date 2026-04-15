import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# APP_VERSION = "v3-2025-04-11"
st.set_page_config(page_title="Inventory Hub", page_icon="📦", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.stApp{background-color:#EEF2F7;}
section[data-testid="stSidebar"]{background-color:#162447!important;}
section[data-testid="stSidebar"] *{color:#E8F0FE!important;}
.stTabs [data-baseweb="tab-list"]{background:#162447;border-radius:0;}
.stTabs [data-baseweb="tab"]{color:#A5B4FC;font-size:13px;}
.stTabs [aria-selected="true"]{color:#fff!important;border-bottom:2px solid #4F46E5!important;}
.stButton>button[kind="primary"]{background:#4F46E5;border:none;color:#fff;}
.stButton>button[kind="primary"]:hover{background:#4338CA;}
</style>
""", unsafe_allow_html=True)

# ── Constants ──
ACTIVOS=['Carhartt','Champion','Duluth Trading','Harley-Davidson Motor Company, Inc.',
         'Kerusso inc.','STUSSY, Inc','Tiltworks','Under Armour','Uniforms','Vortex Optics']
INACTIVOS=['Adidas','CCM Hockey','Dallas Cowboys','Dickies','Fanatics Licensed Sports Group',
           'G&G Outfitters, INC','GFS','Grupo Beta','Jansport','Knigths Apparel','Outerstuff LTD',
           'Profile','Southern Tide','UA-GFS','Upward Sports','Golds Gym','Walls','SLD','Oakley']
NO_VMI=['Duluth Trading','Tiltworks','Vortex Optics']
CMAP_HN={'Regular':'Regulars','VMI':'VMI','Exceso':'Excess','Irregulares':'Irregulars','Obsoleto':'Obsolete'}
CMAP_TLP={}
CMAP_HN_CUSTOMER={}
TLP_ORDER=['Crazy Shirts','Dickies','Grupo Beta','INX PRINTS INC.','Outerstuff LTD','Renfro',
           'Simply Southern','Textiles La Paz','TLP OUTERSTUFF','TLP Vans','Under Armour','Vans','NOVUS','Timberland']

# ── Excel style helpers ──
XNAVY='FF0D2B4E';XNAVY_MID='FF1A4A7A';XNAVY_LIGHT='FF2E6DA4';XNAVY_PALE='FFD6E4F0'
XNAVY_ALT='FFEBF3FB';XWHITE='FFFFFFFF';XGRAY_MID='FFD0D8E4';XACCENT='FF1B6CA8';XSUBTOT='FF2E6DA4';XGRAND='FF0D2B4E'

def brd():
    t=Side(style='thin',color=XGRAY_MID)
    return Border(left=t,right=t,top=t,bottom=t)

def xh(ws,r,c,v,bg=XNAVY,fg=XWHITE,sz=10,b=True,ha='center'):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=b,color=fg,name='Calibri',size=sz)
    cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal=ha,vertical='center',wrap_text=True)
    cell.border=brd();return cell

def xd(ws,r,c,v,bg=None,b=False,ha='right'):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=b,color='FF1A1A2E',name='Calibri',size=10)
    if bg:cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal=ha,vertical='center')
    cell.border=brd()
    if isinstance(v,(int,float)) and v is not None:cell.number_format='#,##0'
    return cell

def xdf(ws,r,c,v,bg=None,hmode=False):
    fg=XWHITE if hmode else ('1E6B3C' if v and v>0 else '8B1A1A' if v and v<0 else '1A1A2E')
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=True,color=fg,name='Calibri',size=10)
    if bg:cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal='right',vertical='center')
    cell.border=brd()
    if isinstance(v,(int,float)) and v is not None:cell.number_format='+#,##0;-#,##0;0'
    return cell

# ── Program classification ──
def apply_program(df):
    df=df.copy();df['Program']=None
    df.loc[df['Box Usage'].astype(str).str.upper().str.contains('BLANK',na=False),'Program']='BLANKS'
    m2=df['Program'].isna()&df['Box Usage'].astype(str).str.upper().str.contains('PRINTED|EMBROIDERY',na=False)
    df.loc[m2,'Program']='PRINTED'
    df.loc[df['Program'].isna()&df['Option6'].isna(),'Program']='BLANKS'
    mo=df['Program'].isna()&df['Option6'].notna()
    df.loc[mo&df['Box\nStatus'].isin(['Packed','Picked']),'Program']='PRINTED'
    df.loc[mo&df['Box\nStatus'].isin(['Inventory','WIP']),'Program']='BLANKS'
    return df

# ── Honduras classification ──
def classify_honduras(carton_df, open_df):
    df=carton_df.copy()
    df['Quantity']=pd.to_numeric(df['Quantity'].astype(str).str.replace(',',''),errors='coerce').fillna(0)
    df['Customer Name']=df['Customer'].astype(str)
    df['Clasificacion']=None
    open_po=set(open_df['PONumber'].astype(str).str.strip().dropna())
    no_vmi=df['Customer Name'].isin(NO_VMI)
    df.loc[df['Is Second'].isin(['Second','Third']),'Clasificacion']='Irregulares'
    df.loc[(df['Order\nType']=='Locker Stock')&~no_vmi&df['Clasificacion'].isna(),'Clasificacion']='VMI'
    df.loc[(df['Customer']==12)&(df['Color\nDescription']=='PFD White')&~no_vmi&df['Clasificacion'].isna(),'Clasificacion']='VMI'
    df.loc[(df['Customer']==81)&df['Style'].astype(str).str.upper().str.endswith('-VMI')&~no_vmi&df['Clasificacion'].isna(),'Clasificacion']='VMI'
    mp=(df['Customer']==81)&df['Style'].astype(str).str.upper().str.contains('VMI-PULL')&~no_vmi
    df.loc[mp&~df['PONumber'].astype(str).str.strip().isin(open_po)&df['Clasificacion'].isna(),'Clasificacion']='VMI'
    df.loc[mp&df['PONumber'].astype(str).str.strip().isin(open_po)&df['Clasificacion'].isna(),'Clasificacion']='Regular'
    p1=df['PONumber'].astype(str).str.strip();p2=df['PONumbers'].astype(str).str.strip()
    df.loc[df['Clasificacion'].isna()&(p1.isin(open_po)|p2.isin(open_po)),'Clasificacion']='Regular'
    df.loc[df['Clasificacion'].isna()&df['Box Tag'].astype(str).str.contains('Excess',na=False),'Clasificacion']='Exceso'
    df.loc[df['Clasificacion'].isna()&(df['Box Tag']=='Obsolete'),'Clasificacion']='Obsoleto'
    df['Create Date']=pd.to_datetime(df['Create Date'],errors='coerce')
    cutoff=datetime.now().replace(year=datetime.now().year-1)
    ms=df['Order Status'].isin(['Complete','Void'])
    df.loc[df['Clasificacion'].isna()&ms&(df['Create Date']<=cutoff),'Clasificacion']='Obsoleto'
    df.loc[df['Clasificacion'].isna()&ms&(df['Create Date']>cutoff),'Clasificacion']='Exceso'
    df.loc[df['Clasificacion'].isna()&(df['Create Date']<=cutoff),'Clasificacion']='Obsoleto'
    df.loc[df['Clasificacion'].isna()&(df['Create Date']>cutoff),'Clasificacion']='Exceso'
    cut=df['Box Usage'].astype(str).str.upper().str.contains('CUT',na=False)
    cut_info=(cut.sum(),int(df.loc[cut,'Quantity'].sum()))
    df=df[~cut].copy()
    bs=df['Box\nStatus'];cl=df['Clasificacion']
    df['Type']=''
    df.loc[(cl=='Regular')&bs.isin(['Packed','Picked']),'Type']='Finished Goods'
    df.loc[(cl=='Regular')&bs.isin(['Inventory','WIP']),'Type']='Wip'
    df.loc[cl.isin(['VMI','Irregulares','Obsoleto','Exceso']),'Type']='Finished Goods'
    # Rename Regular Wip to avoid duplication in tables
    df.loc[(df['Clasificacion']=='Regular')&(df['Type']=='Wip'),'Clasificacion']='Regular Wip'
    df=apply_program(df)
    df['Year']=df['Create Date'].dt.year.astype('Int64')
    return df,cut_info

# ── TLP classification ──
def classify_tlp(carton_df):
    df=carton_df.copy()
    df['Quantity']=pd.to_numeric(df['Quantity'].astype(str).str.replace(',',''),errors='coerce').fillna(0)
    df['Customer Name']=df['Customer'].astype(str)
    df['Clasificacion']=None
    df.loc[df['Is Second'].isin(['Second','Third']),'Clasificacion']='TLP Irregulars'
    df.loc[df['Clasificacion'].isna()&(df['Box Tag']=='Blanks Excess'),'Clasificacion']='TLP Blanks Excess'
    df.loc[df['Clasificacion'].isna()&(df['Box Tag']=='Printed Excess'),'Clasificacion']='TLP Printed Excess'
    df.loc[df['Clasificacion'].isna()&df['Box\nStatus'].isin(['Packed','Picked']),'Clasificacion']='TLP sin clasificacion'
    df.loc[df['Clasificacion'].isna(),'Clasificacion']='Wip'
    # Standardize TLP classification names
    df['Clasificacion'] = df['Clasificacion'].replace({
        'TLP Irregulars': 'Irregulares',
        'TLP Blanks Excess': 'Exceso Blanks',
        'TLP Printed Excess': 'Exceso Printed',
        'TLP sin clasificacion': 'Sin Clasificacion',
    })
    df=apply_program(df)
    df['Create Date']=pd.to_datetime(df['Create Date'],errors='coerce')
    df['Year']=df['Create Date'].dt.year.astype('Int64')
    return df

# ── Excel pivot writer ──
def write_pivot_sheet(ws,df,inv_cols,activos,inactivos):
    df=df.copy();df['Clas_Col']=df['Clasificacion'].map(CMAP_HN)
    p=df.pivot_table(index='Customer Name',columns='Clas_Col',values='Quantity',aggfunc='sum',fill_value=0)
    for c in inv_cols:
        if c not in p.columns:p[c]=0
    p=p[inv_cols];p['Grand Total']=p.sum(axis=1)
    cols=inv_cols+['Grand Total']
    ws.merge_cells('C1:H1');ws['C1']='Inventory Type'
    ws['C1'].font=Font(bold=True,name='Calibri',size=10,color=XNAVY)
    ws['C1'].alignment=Alignment(horizontal='center')
    for ci,h in enumerate(['Clientes','Style CustomerName']+cols,1):xh(ws,2,ci,h)
    row=3
    for gn,gl in [('Activos',activos),('Inactivos',inactivos)]:
        clients=[c for c in gl if c in p.index]
        if not clients:continue
        first=True
        for i,cli in enumerate(clients):
            r=p.loc[cli];fill=XNAVY_ALT if i%2==0 else None
            c1=ws.cell(row=row,column=1,value=gn if first else '')
            c1.font=Font(bold=True,color=XWHITE,name='Calibri',size=10)
            c1.fill=PatternFill('solid',start_color=XACCENT)
            c1.alignment=Alignment(horizontal='center',vertical='center')
            c1.border=brd();first=False
            xd(ws,row,2,cli,bg=fill,ha='left')
            for ci,col in enumerate(cols,3):xd(ws,row,ci,int(r[col]) if r[col]!=0 else None,bg=fill)
            row+=1
        st_row=p.loc[clients].sum()
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        xh(ws,row,1,f'{gn} Total',XSUBTOT)
        for ci,col in enumerate(cols,3):xh(ws,row,ci,int(st_row[col]) if st_row[col]!=0 else None,XSUBTOT)
        row+=1
    gt=p.sum()
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    xh(ws,row,1,'Grand Total',XGRAND)
    for ci,col in enumerate(cols,3):xh(ws,row,ci,int(gt[col]) if gt[col]!=0 else None,XGRAND)
    ws.column_dimensions['A'].width=14;ws.column_dimensions['B'].width=38
    for ci in range(3,3+len(cols)):ws.column_dimensions[get_column_letter(ci)].width=15
    ws.freeze_panes='A3'

# ── Honduras Excel ──
def build_excel_hn(df,wk_prev=None,wk_label='WK13'):
    inv_cols=['Regulars','VMI','Excess','Irregulars','Obsolete']
    years=sorted(df['Year'].dropna().unique())
    wb=Workbook();wb.remove(wb.active)

    write_pivot_sheet(wb.create_sheet('Resumen Total'),df,inv_cols,ACTIVOS,INACTIVOS)
    write_pivot_sheet(wb.create_sheet('Finished Goods'),df[df['Type']=='Finished Goods'],inv_cols,ACTIVOS,INACTIVOS)
    write_pivot_sheet(wb.create_sheet('Wip'),df[df['Type']=='Wip'],inv_cols,ACTIVOS,INACTIVOS)

    # Antigüedad
    ws_age=wb.create_sheet('Antigüedad')
    for label,d in [('Finished Goods',df[df['Type']=='Finished Goods']),('Wip',df[df['Type']=='Wip'])]:
        p=d.pivot_table(index='Customer Name',columns='Year',values='Quantity',aggfunc='sum',fill_value=0)
        for yr in years:
            if yr not in p.columns:p[yr]=0
        p=p[[yr for yr in years]];p['Grand Total']=p.sum(axis=1)
        sr=ws_age.max_row+1 if ws_age.max_row>1 else 1
        ncols=2+len(years)+1
        ws_age.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=ncols)
        c=ws_age.cell(row=sr,column=1,value=f'  {label}')
        c.font=Font(bold=True,color=XWHITE,name='Calibri',size=12)
        c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
        c.border=brd();ws_age.row_dimensions[sr].height=22;sr+=1
        ws_age.merge_cells(start_row=sr,start_column=3,end_row=sr,end_column=2+len(years))
        xh(ws_age,sr,1,'Clientes',XNAVY_MID);xh(ws_age,sr,2,'Customer Name',XNAVY_MID)
        xh(ws_age,sr,3,'Year of Creation',XNAVY_MID)
        for ci in range(4,2+len(years)+1):
            c=ws_age.cell(row=sr,column=ci);c.fill=PatternFill('solid',start_color=XNAVY_MID);c.border=brd()
        xh(ws_age,sr,2+len(years)+1,'Grand Total',XNAVY_MID);sr+=1
        xh(ws_age,sr,1,'',XNAVY_LIGHT);xh(ws_age,sr,2,'',XNAVY_LIGHT)
        for yi,yr in enumerate(years):xh(ws_age,sr,3+yi,int(yr),XNAVY_LIGHT)
        xh(ws_age,sr,3+len(years),'Grand Total',XNAVY_LIGHT);sr+=1
        for gn,gl in [('Activos',ACTIVOS),('Inactivos',INACTIVOS)]:
            clients=[c for c in gl if c in p.index]
            if not clients:continue
            first=True
            for i,cli in enumerate(clients):
                r=p.loc[cli];rb=XNAVY_ALT if i%2==0 else None
                c1=ws_age.cell(row=sr,column=1,value=gn if first else '')
                c1.font=Font(bold=True,color=XWHITE,name='Calibri',size=10)
                c1.fill=PatternFill('solid',start_color=XACCENT)
                c1.alignment=Alignment(horizontal='center',vertical='center');c1.border=brd();first=False
                xd(ws_age,sr,2,cli,bg=rb,ha='left')
                for yi,yr in enumerate(years):xd(ws_age,sr,3+yi,int(r[yr]) if r[yr]!=0 else None,bg=rb)
                xd(ws_age,sr,3+len(years),int(r['Grand Total']),bg=XNAVY_PALE,b=True);sr+=1
            st2=p.loc[clients].sum()
            ws_age.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=2)
            xh(ws_age,sr,1,f'{gn} Total',XSUBTOT)
            for yi,yr in enumerate(years):xh(ws_age,sr,3+yi,int(st2[yr]) if st2[yr]!=0 else None,XSUBTOT)
            xh(ws_age,sr,3+len(years),int(st2['Grand Total']),XSUBTOT);sr+=1
        gt=p.sum()
        ws_age.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=2)
        xh(ws_age,sr,1,'Grand Total',XGRAND)
        for yi,yr in enumerate(years):xh(ws_age,sr,3+yi,int(gt[yr]) if gt[yr]!=0 else None,XGRAND)
        xh(ws_age,sr,3+len(years),int(gt['Grand Total']),XGRAND)

    ws_age.column_dimensions['A'].width=14;ws_age.column_dimensions['B'].width=36
    for ci in range(3,3+len(years)+1):ws_age.column_dimensions[get_column_letter(ci)].width=10
    ws_age.column_dimensions[get_column_letter(3+len(years))].width=13;ws_age.freeze_panes='C4'

    # Damage Severity
    ws_dmg=wb.create_sheet('Damage Severity')
    d=df[df['Clasificacion']=='Irregulares'].copy()
    d['DS_Group']=d['Damage Severity'].apply(lambda x:'0, 1' if x in [0,1] else '2+')
    p=d.pivot_table(index='Customer Name',columns='DS_Group',values='Quantity',aggfunc='sum',fill_value=0)
    for c in ['0, 1','2+']:
        if c not in p.columns:p[c]=0
    p=p[['0, 1','2+']];p['Grand Total']=p.sum(axis=1)
    p=p.sort_values('Grand Total',ascending=False)
    ws_dmg.merge_cells('A1:E1')
    c=ws_dmg.cell(row=1,column=1,value='  Damage Severity — Irregulars Summary')
    c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
    c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=brd();ws_dmg.row_dimensions[1].height=26
    ws_dmg.merge_cells('C2:D2')
    xh(ws_dmg,2,1,'Inventory Type',XNAVY_MID);xh(ws_dmg,2,2,'Style CustomerName',XNAVY_MID)
    xh(ws_dmg,2,3,'Damage Severity',XNAVY_MID)
    ws_dmg.cell(row=2,column=4).fill=PatternFill('solid',start_color=XNAVY_MID);ws_dmg.cell(row=2,column=4).border=brd()
    xh(ws_dmg,2,5,'Grand Total',XNAVY_MID)
    xh(ws_dmg,3,1,'',XNAVY_LIGHT);xh(ws_dmg,3,2,'',XNAVY_LIGHT)
    xh(ws_dmg,3,3,'0, 1',XNAVY_LIGHT);xh(ws_dmg,3,4,'2 +',XNAVY_LIGHT);xh(ws_dmg,3,5,'Grand Total',XNAVY_LIGHT)
    row=4;first=True
    for i,(cli,r) in enumerate(p.iterrows()):
        fill=XNAVY_ALT if i%2==0 else None
        c1=ws_dmg.cell(row=row,column=1,value='Irregulars' if first else '')
        c1.font=Font(bold=True,color=XWHITE,name='Calibri',size=10)
        c1.fill=PatternFill('solid',start_color=XACCENT)
        c1.alignment=Alignment(horizontal='center',vertical='center');c1.border=brd();first=False
        xd(ws_dmg,row,2,cli,bg=fill,ha='left')
        xd(ws_dmg,row,3,int(r['0, 1']) if r['0, 1']!=0 else None,bg=fill)
        xd(ws_dmg,row,4,int(r['2+']) if r['2+']!=0 else None,bg=fill)
        xd(ws_dmg,row,5,int(r['Grand Total']),bg=fill,b=True);row+=1
    gt=p.sum()
    ws_dmg.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
    xh(ws_dmg,row,1,'Grand Total',XGRAND)
    xh(ws_dmg,row,3,int(gt['0, 1']) if gt['0, 1']!=0 else None,XGRAND)
    xh(ws_dmg,row,4,int(gt['2+']) if gt['2+']!=0 else None,XGRAND)
    xh(ws_dmg,row,5,int(gt['Grand Total']),XGRAND)
    for col,w in zip('ABCDE',[16,28,14,14,14]):ws_dmg.column_dimensions[col].width=w
    ws_dmg.freeze_panes='C4'

    # Program
    ws_prog=wb.create_sheet('Program')
    dp=df[df['Type']=='Finished Goods'].copy()
    pp=dp.pivot_table(index='Clasificacion',columns='Program',values='Quantity',aggfunc='sum',fill_value=0)
    for c in ['BLANKS','PRINTED']:
        if c not in pp.columns:pp[c]=0
    pp=pp[['BLANKS','PRINTED']];pp['Grand Total']=pp.sum(axis=1)
    order=['Regular','VMI','Exceso','Irregulares','Obsoleto']
    pp=pp.reindex([c for c in order if c in pp.index]+[c for c in pp.index if c not in order])
    ws_prog.merge_cells('A1:E1')
    c=ws_prog.cell(row=1,column=1,value='  Program Summary — BLANKS vs PRINTED  (Finished Goods)')
    c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
    c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=brd();ws_prog.row_dimensions[1].height=26
    ws_prog.merge_cells('B2:C2')
    xh(ws_prog,2,1,'Inventory Type',XNAVY_MID);xh(ws_prog,2,2,'Program',XNAVY_MID)
    ws_prog.cell(row=2,column=3).fill=PatternFill('solid',start_color=XNAVY_MID);ws_prog.cell(row=2,column=3).border=brd()
    xh(ws_prog,2,4,'Grand Total',XNAVY_MID)
    xh(ws_prog,3,1,'',XNAVY_LIGHT);xh(ws_prog,3,2,'BLANKS',XNAVY_LIGHT)
    xh(ws_prog,3,3,'PRINTED',XNAVY_LIGHT);xh(ws_prog,3,4,'Grand Total',XNAVY_LIGHT)
    for i,(clas,r) in enumerate(pp.iterrows()):
        fill=XNAVY_ALT if i%2==0 else None
        xd(ws_prog,4+i,1,clas,bg=fill,ha='left')
        xd(ws_prog,4+i,2,int(r['BLANKS']) if r['BLANKS']!=0 else None,bg=fill)
        xd(ws_prog,4+i,3,int(r['PRINTED']) if r['PRINTED']!=0 else None,bg=fill)
        xd(ws_prog,4+i,4,int(r['Grand Total']),bg=fill,b=True)
    last=4+len(pp);gt=pp.sum()
    xh(ws_prog,last,1,'Grand Total',XGRAND,ha='left')
    xh(ws_prog,last,2,int(gt['BLANKS']) if gt['BLANKS']!=0 else None,XGRAND)
    xh(ws_prog,last,3,int(gt['PRINTED']) if gt['PRINTED']!=0 else None,XGRAND)
    xh(ws_prog,last,4,int(gt['Grand Total']),XGRAND)
    for col,w in zip('ABCD',[26,16,16,16]):ws_prog.column_dimensions[col].width=w
    ws_prog.freeze_panes='A4'

    # Comparativo
    if wk_prev:
        inv_types=['Regulars','VMI','Excess','Irregulars','Obsolete','Liability']
        df2=df.copy();df2['Clas_Col']=df2['Clasificacion'].map(CMAP_HN)
        wk13_p=df2[df2['Type']=='Finished Goods'].pivot_table(index='Customer Name',columns='Clas_Col',values='Quantity',aggfunc='sum',fill_value=0)
        for t in inv_types:
            if t not in wk13_p.columns:wk13_p[t]=0
        wk13_p['Liability']=0
        ws_comp=wb.create_sheet('Comparativo')
        ncols=2+len(inv_types)*3+3
        ws_comp.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
        c=ws_comp.cell(row=1,column=1,value=f'  Inventory Comparison — {wk_label} vs WK Anterior  (Finished Goods)')
        c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
        c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
        c.border=brd();ws_comp.row_dimensions[1].height=26
        xh(ws_comp,2,1,'Clientes',XNAVY_MID);xh(ws_comp,2,2,'Customer Name',XNAVY_MID)
        col=3
        for t in inv_types:
            ws_comp.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+2)
            xh(ws_comp,2,col,t,XNAVY_MID);col+=3
        ws_comp.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+2)
        xh(ws_comp,2,col,'Grand Total',XNAVY)
        xh(ws_comp,3,1,'',XNAVY_LIGHT);xh(ws_comp,3,2,'',XNAVY_LIGHT)
        col=3
        for t in inv_types:
            xh(ws_comp,3,col,'WK Ant.',XNAVY_LIGHT);xh(ws_comp,3,col+1,wk_label,XNAVY_LIGHT);xh(ws_comp,3,col+2,'Diff',XNAVY_LIGHT);col+=3
        xh(ws_comp,3,col,'WK Ant.',XNAVY);xh(ws_comp,3,col+1,wk_label,XNAVY);xh(ws_comp,3,col+2,'Diff',XNAVY)
        row=4;tot12={t:0 for t in inv_types};tot13={t:0 for t in inv_types}
        for gn,gl in [('Activos',ACTIVOS),('Inactivos',INACTIVOS)]:
            first=True
            for i,cli in enumerate(gl):
                w12=wk_prev.get(cli,{t:0 for t in inv_types})
                w13={t:int(wk13_p.loc[cli,t]) if cli in wk13_p.index and t in wk13_p.columns else 0 for t in inv_types}
                fill=XNAVY_ALT if i%2==0 else None
                c1=ws_comp.cell(row=row,column=1,value=gn if first else '')
                c1.font=Font(bold=True,color=XWHITE,name='Calibri',size=10)
                c1.fill=PatternFill('solid',start_color=XACCENT)
                c1.alignment=Alignment(horizontal='center',vertical='center');c1.border=brd();first=False
                xd(ws_comp,row,2,cli,bg=fill,ha='left')
                col=3;gt12=gt13=0
                for t in inv_types:
                    v12=w12[t];v13=w13[t];diff=v13-v12
                    xd(ws_comp,row,col,v12 if v12 else None,bg=fill)
                    xd(ws_comp,row,col+1,v13 if v13 else None,bg=fill)
                    xdf(ws_comp,row,col+2,diff if diff!=0 else None,bg=fill)
                    gt12+=v12;gt13+=v13;tot12[t]+=v12;tot13[t]+=v13;col+=3
                xd(ws_comp,row,col,gt12 if gt12 else None,bg=fill,b=True)
                xd(ws_comp,row,col+1,gt13 if gt13 else None,bg=fill,b=True)
                xdf(ws_comp,row,col+2,(gt13-gt12) if gt13-gt12!=0 else None,bg=fill);row+=1
            s12={t:sum(wk_prev.get(c,{t:0 for t in inv_types})[t] for c in gl) for t in inv_types}
            s13={t:sum(int(wk13_p.loc[c,t]) if c in wk13_p.index and t in wk13_p.columns else 0 for c in gl) for t in inv_types}
            ws_comp.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
            xh(ws_comp,row,1,f'{gn} Total',XSUBTOT);col=3;sgt12=sgt13=0
            for t in inv_types:
                xh(ws_comp,row,col,s12[t] if s12[t] else None,XSUBTOT)
                xh(ws_comp,row,col+1,s13[t] if s13[t] else None,XSUBTOT)
                xdf(ws_comp,row,col+2,(s13[t]-s12[t]) if s13[t]-s12[t]!=0 else None,bg=XSUBTOT,hmode=True)
                sgt12+=s12[t];sgt13+=s13[t];col+=3
            xh(ws_comp,row,col,sgt12,XSUBTOT);xh(ws_comp,row,col+1,sgt13,XSUBTOT)
            xdf(ws_comp,row,col+2,sgt13-sgt12,bg=XSUBTOT,hmode=True);row+=1
        ws_comp.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        xh(ws_comp,row,1,'Grand Total',XGRAND);col=3;ggt12=ggt13=0
        for t in inv_types:
            xh(ws_comp,row,col,tot12[t] if tot12[t] else None,XGRAND)
            xh(ws_comp,row,col+1,tot13[t] if tot13[t] else None,XGRAND)
            xdf(ws_comp,row,col+2,(tot13[t]-tot12[t]) if tot13[t]-tot12[t]!=0 else None,bg=XGRAND,hmode=True)
            ggt12+=tot12[t];ggt13+=tot13[t];col+=3
        xh(ws_comp,row,col,ggt12,XGRAND);xh(ws_comp,row,col+1,ggt13,XGRAND)
        xdf(ws_comp,row,col+2,ggt13-ggt12,bg=XGRAND,hmode=True)
        ws_comp.column_dimensions['A'].width=12;ws_comp.column_dimensions['B'].width=36
        col=3
        for _ in inv_types:
            ws_comp.column_dimensions[get_column_letter(col)].width=11
            ws_comp.column_dimensions[get_column_letter(col+1)].width=11
            ws_comp.column_dimensions[get_column_letter(col+2)].width=10;col+=3
        for i in range(3):ws_comp.column_dimensions[get_column_letter(col+i)].width=12
        ws_comp.freeze_panes='C4'

    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf

# ── TLP Excel ──
def build_excel_tlp(df,wk_prev=None,wk_label='WK13'):
    inv_tlp=['TLP Irregulars','TLP Printed Excess','TLP sin clasificacion','TLP Blanks Excess','Wip']
    years=sorted(df['Year'].dropna().unique())
    all_clients=[c for c in TLP_ORDER if c in df['Customer Name'].unique()]+[c for c in df['Customer Name'].unique() if c not in TLP_ORDER]
    wb=Workbook();wb.remove(wb.active)

    # Resumen Total
    ws1=wb.create_sheet('Resumen Total')
    p=df.pivot_table(index='Customer Name',columns='Clasificacion',values='Quantity',aggfunc='sum',fill_value=0)
    for t in inv_tlp:
        if t not in p.columns:p[t]=0
    p['Grand Total']=p[inv_tlp].sum(axis=1)
    ws1.merge_cells(start_row=1,start_column=3,end_row=1,end_column=2+len(inv_tlp))
    ws1['C1']='Inventory Type';ws1['C1'].font=Font(bold=True,name='Calibri',size=10,color=XNAVY)
    ws1['C1'].alignment=Alignment(horizontal='center')
    for ci,h in enumerate(['Cliente','Style CustomerName']+inv_tlp+['Grand Total'],1):xh(ws1,2,ci,h)
    for i,cli in enumerate(all_clients):
        if cli not in p.index:continue
        r=p.loc[cli];fill=XNAVY_ALT if i%2==0 else None
        ws1.cell(row=3+i,column=1,value='').border=brd()
        xd(ws1,3+i,2,cli,bg=fill,ha='left')
        for ci,col in enumerate(inv_tlp+['Grand Total'],3):xd(ws1,3+i,ci,int(r[col]) if r[col]!=0 else None,bg=fill)
    last=3+len(all_clients);gt=p[inv_tlp+['Grand Total']].sum()
    ws1.merge_cells(start_row=last,start_column=1,end_row=last,end_column=2);xh(ws1,last,1,'Grand Total',XGRAND)
    for ci,col in enumerate(inv_tlp+['Grand Total'],3):xh(ws1,last,ci,int(gt[col]) if gt[col]!=0 else None,XGRAND)
    ws1.column_dimensions['A'].width=5;ws1.column_dimensions['B'].width=22
    for ci in range(3,3+len(inv_tlp)+2):ws1.column_dimensions[get_column_letter(ci)].width=22
    ws1.freeze_panes='B3'

    # Antigüedad
    ws_age=wb.create_sheet('Antigüedad')
    p2=df.pivot_table(index='Customer Name',columns='Year',values='Quantity',aggfunc='sum',fill_value=0)
    for yr in years:
        if yr not in p2.columns:p2[yr]=0
    p2=p2[[yr for yr in years]];p2['Grand Total']=p2.sum(axis=1)
    ncols=2+len(years)+1
    ws_age.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws_age.cell(row=1,column=1,value='  Antigüedad del Inventario TLP')
    c.font=Font(bold=True,color=XWHITE,name='Calibri',size=12)
    c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=brd();ws_age.row_dimensions[1].height=22
    ws_age.merge_cells(start_row=2,start_column=3,end_row=2,end_column=2+len(years))
    xh(ws_age,2,1,'',XNAVY_MID);xh(ws_age,2,2,'Customer Name',XNAVY_MID);xh(ws_age,2,3,'Year of Creation',XNAVY_MID)
    for ci in range(4,2+len(years)+1):
        c=ws_age.cell(row=2,column=ci);c.fill=PatternFill('solid',start_color=XNAVY_MID);c.border=brd()
    xh(ws_age,2,2+len(years)+1,'Grand Total',XNAVY_MID)
    xh(ws_age,3,1,'',XNAVY_LIGHT);xh(ws_age,3,2,'',XNAVY_LIGHT)
    for yi,yr in enumerate(years):xh(ws_age,3,3+yi,int(yr),XNAVY_LIGHT)
    xh(ws_age,3,3+len(years),'Grand Total',XNAVY_LIGHT)
    for i,cli in enumerate(all_clients):
        if cli not in p2.index:continue
        r=p2.loc[cli];rb=XNAVY_ALT if i%2==0 else None
        ws_age.cell(row=4+i,column=1,value='').border=brd()
        xd(ws_age,4+i,2,cli,bg=rb,ha='left')
        for yi,yr in enumerate(years):xd(ws_age,4+i,3+yi,int(r[yr]) if r[yr]!=0 else None,bg=rb)
        xd(ws_age,4+i,3+len(years),int(r['Grand Total']),bg=XNAVY_PALE,b=True)
    last2=4+len(all_clients);gt2=p2.sum()
    ws_age.merge_cells(start_row=last2,start_column=1,end_row=last2,end_column=2);xh(ws_age,last2,1,'Grand Total',XGRAND)
    for yi,yr in enumerate(years):xh(ws_age,last2,3+yi,int(gt2[yr]) if gt2[yr]!=0 else None,XGRAND)
    xh(ws_age,last2,3+len(years),int(gt2['Grand Total']),XGRAND)
    ws_age.column_dimensions['A'].width=5;ws_age.column_dimensions['B'].width=22
    for ci in range(3,3+len(years)+1):ws_age.column_dimensions[get_column_letter(ci)].width=10
    ws_age.column_dimensions[get_column_letter(3+len(years))].width=13;ws_age.freeze_panes='C4'

    # Damage Severity
    ws_dmg=wb.create_sheet('Damage Severity')
    dd=df[df['Clasificacion']=='TLP Irregulars'].copy()
    dd['DS_Group']=dd['Damage Severity'].apply(lambda x:'0, 1' if x in [0,1] else '2+')
    pd3=dd.pivot_table(index='Customer Name',columns='DS_Group',values='Quantity',aggfunc='sum',fill_value=0)
    for c in ['0, 1','2+']:
        if c not in pd3.columns:pd3[c]=0
    pd3=pd3[['0, 1','2+']];pd3['Grand Total']=pd3.sum(axis=1)
    ordered=[c for c in TLP_ORDER if c in pd3.index]+[c for c in pd3.index if c not in TLP_ORDER]
    pd3=pd3.reindex(ordered)
    ws_dmg.merge_cells('A1:E1')
    c=ws_dmg.cell(row=1,column=1,value='  Damage Severity — TLP Irregulars Summary')
    c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
    c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=brd();ws_dmg.row_dimensions[1].height=26
    ws_dmg.merge_cells('C2:D2')
    xh(ws_dmg,2,1,'Inventory Type',XNAVY_MID);xh(ws_dmg,2,2,'Style CustomerName',XNAVY_MID)
    xh(ws_dmg,2,3,'Damage Severity',XNAVY_MID)
    ws_dmg.cell(row=2,column=4).fill=PatternFill('solid',start_color=XNAVY_MID);ws_dmg.cell(row=2,column=4).border=brd()
    xh(ws_dmg,2,5,'Grand Total',XNAVY_MID)
    xh(ws_dmg,3,1,'',XNAVY_LIGHT);xh(ws_dmg,3,2,'',XNAVY_LIGHT)
    xh(ws_dmg,3,3,'0, 1',XNAVY_LIGHT);xh(ws_dmg,3,4,'2 +',XNAVY_LIGHT);xh(ws_dmg,3,5,'Grand Total',XNAVY_LIGHT)
    row=4;first=True
    for i,(cli,r) in enumerate(pd3.iterrows()):
        fill=XNAVY_ALT if i%2==0 else None
        c1=ws_dmg.cell(row=row,column=1,value='TLP Irregulars' if first else '')
        c1.font=Font(bold=True,color=XWHITE,name='Calibri',size=10)
        c1.fill=PatternFill('solid',start_color=XACCENT)
        c1.alignment=Alignment(horizontal='center',vertical='center');c1.border=brd();first=False
        xd(ws_dmg,row,2,cli,bg=fill,ha='left')
        xd(ws_dmg,row,3,int(r['0, 1']) if r['0, 1']!=0 else None,bg=fill)
        xd(ws_dmg,row,4,int(r['2+']) if r['2+']!=0 else None,bg=fill)
        xd(ws_dmg,row,5,int(r['Grand Total']),bg=fill,b=True);row+=1
    gt3=pd3.sum()
    ws_dmg.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2);xh(ws_dmg,row,1,'Grand Total',XGRAND)
    xh(ws_dmg,row,3,int(gt3['0, 1']) if gt3['0, 1']!=0 else None,XGRAND)
    xh(ws_dmg,row,4,int(gt3['2+']) if gt3['2+']!=0 else None,XGRAND)
    xh(ws_dmg,row,5,int(gt3['Grand Total']),XGRAND)
    for col,w in zip('ABCDE',[16,28,14,14,14]):ws_dmg.column_dimensions[col].width=w
    ws_dmg.freeze_panes='C4'

    # Program
    ws_prog=wb.create_sheet('Program')
    dp=df[df['Clasificacion']!='Wip'].copy()
    pp=dp.pivot_table(index='Clasificacion',columns='Program',values='Quantity',aggfunc='sum',fill_value=0)
    for c in ['BLANKS','PRINTED']:
        if c not in pp.columns:pp[c]=0
    pp=pp[['BLANKS','PRINTED']];pp['Grand Total']=pp.sum(axis=1)
    order2=['TLP Irregulars','TLP Printed Excess','TLP sin clasificacion','TLP Blanks Excess']
    pp=pp.reindex([c for c in order2 if c in pp.index]+[c for c in pp.index if c not in order2])
    ws_prog.merge_cells('A1:E1')
    c=ws_prog.cell(row=1,column=1,value='  Program Summary — BLANKS vs PRINTED  (Finished Goods)')
    c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
    c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
    c.border=brd();ws_prog.row_dimensions[1].height=26
    ws_prog.merge_cells('B2:C2')
    xh(ws_prog,2,1,'Inventory Type',XNAVY_MID);xh(ws_prog,2,2,'Program',XNAVY_MID)
    ws_prog.cell(row=2,column=3).fill=PatternFill('solid',start_color=XNAVY_MID);ws_prog.cell(row=2,column=3).border=brd()
    xh(ws_prog,2,4,'Grand Total',XNAVY_MID)
    xh(ws_prog,3,1,'',XNAVY_LIGHT);xh(ws_prog,3,2,'BLANKS',XNAVY_LIGHT)
    xh(ws_prog,3,3,'PRINTED',XNAVY_LIGHT);xh(ws_prog,3,4,'Grand Total',XNAVY_LIGHT)
    for i,(clas,r) in enumerate(pp.iterrows()):
        fill=XNAVY_ALT if i%2==0 else None
        xd(ws_prog,4+i,1,clas,bg=fill,ha='left')
        xd(ws_prog,4+i,2,int(r['BLANKS']) if r['BLANKS']!=0 else None,bg=fill)
        xd(ws_prog,4+i,3,int(r['PRINTED']) if r['PRINTED']!=0 else None,bg=fill)
        xd(ws_prog,4+i,4,int(r['Grand Total']),bg=fill,b=True)
    last3=4+len(pp);gt4=pp.sum()
    xh(ws_prog,last3,1,'Grand Total',XGRAND,ha='left')
    xh(ws_prog,last3,2,int(gt4['BLANKS']) if gt4['BLANKS']!=0 else None,XGRAND)
    xh(ws_prog,last3,3,int(gt4['PRINTED']) if gt4['PRINTED']!=0 else None,XGRAND)
    xh(ws_prog,last3,4,int(gt4['Grand Total']),XGRAND)
    for col,w in zip('ABCD',[26,16,16,16]):ws_prog.column_dimensions[col].width=w
    ws_prog.freeze_panes='A4'

    # Comparativo TLP
    if wk_prev:
        inv_comp=['TLP Irregulars','TLP Printed Excess','TLP sin clasificacion','TLP Blanks Excess']
        wk13_p=df.pivot_table(index='Customer Name',columns='Clasificacion',values='Quantity',aggfunc='sum',fill_value=0)
        for t in inv_comp:
            if t not in wk13_p.columns:wk13_p[t]=0
        ws_comp=wb.create_sheet('Comparativo')
        ncols=2+len(inv_comp)*3+3
        ws_comp.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
        c=ws_comp.cell(row=1,column=1,value=f'  Inventory Comparison TLP — {wk_label} vs WK Anterior  (sin Wip)')
        c.font=Font(bold=True,color=XWHITE,name='Calibri',size=13)
        c.fill=PatternFill('solid',start_color=XNAVY);c.alignment=Alignment(horizontal='left',vertical='center')
        c.border=brd();ws_comp.row_dimensions[1].height=26
        xh(ws_comp,2,1,'',XNAVY_MID);xh(ws_comp,2,2,'Customer Name',XNAVY_MID)
        col=3
        for t in inv_comp:
            ws_comp.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+2)
            xh(ws_comp,2,col,t,XNAVY_MID);col+=3
        ws_comp.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+2)
        xh(ws_comp,2,col,'Grand Total',XNAVY)
        xh(ws_comp,3,1,'',XNAVY_LIGHT);xh(ws_comp,3,2,'',XNAVY_LIGHT)
        col=3
        for t in inv_comp:
            xh(ws_comp,3,col,'WK Ant.',XNAVY_LIGHT);xh(ws_comp,3,col+1,wk_label,XNAVY_LIGHT);xh(ws_comp,3,col+2,'Diff',XNAVY_LIGHT);col+=3
        xh(ws_comp,3,col,'WK Ant.',XNAVY);xh(ws_comp,3,col+1,wk_label,XNAVY);xh(ws_comp,3,col+2,'Diff',XNAVY)
        row=4;tot12={t:0 for t in inv_comp};tot13={t:0 for t in inv_comp}
        for i,cli in enumerate(all_clients):
            w12=wk_prev.get(cli,{t:0 for t in inv_comp})
            w13={t:int(wk13_p.loc[cli,t]) if cli in wk13_p.index and t in wk13_p.columns else 0 for t in inv_comp}
            fill=XNAVY_ALT if i%2==0 else None
            ws_comp.cell(row=row,column=1,value='').border=brd()
            xd(ws_comp,row,2,cli,bg=fill,ha='left')
            col=3;gt12=gt13=0
            for t in inv_comp:
                v12=w12[t];v13=w13[t];diff=v13-v12
                xd(ws_comp,row,col,v12 if v12 else None,bg=fill)
                xd(ws_comp,row,col+1,v13 if v13 else None,bg=fill)
                xdf(ws_comp,row,col+2,diff if diff!=0 else None,bg=fill)
                gt12+=v12;gt13+=v13;tot12[t]+=v12;tot13[t]+=v13;col+=3
            xd(ws_comp,row,col,gt12 if gt12 else None,bg=fill,b=True)
            xd(ws_comp,row,col+1,gt13 if gt13 else None,bg=fill,b=True)
            xdf(ws_comp,row,col+2,(gt13-gt12) if gt13-gt12!=0 else None,bg=fill);row+=1
        ws_comp.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        xh(ws_comp,row,1,'Grand Total',XGRAND);col=3;ggt12=ggt13=0
        for t in inv_comp:
            xh(ws_comp,row,col,tot12[t] if tot12[t] else None,XGRAND)
            xh(ws_comp,row,col+1,tot13[t] if tot13[t] else None,XGRAND)
            xdf(ws_comp,row,col+2,(tot13[t]-tot12[t]) if tot13[t]-tot12[t]!=0 else None,bg=XGRAND,hmode=True)
            ggt12+=tot12[t];ggt13+=tot13[t];col+=3
        xh(ws_comp,row,col,ggt12,XGRAND);xh(ws_comp,row,col+1,ggt13,XGRAND)
        xdf(ws_comp,row,col+2,ggt13-ggt12,bg=XGRAND,hmode=True)
        ws_comp.column_dimensions['A'].width=5;ws_comp.column_dimensions['B'].width=22
        col=3
        for _ in inv_comp:
            ws_comp.column_dimensions[get_column_letter(col)].width=20
            ws_comp.column_dimensions[get_column_letter(col+1)].width=20
            ws_comp.column_dimensions[get_column_letter(col+2)].width=12;col+=3
        for i in range(3):ws_comp.column_dimensions[get_column_letter(col+i)].width=14
        ws_comp.freeze_panes='C4'

    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf

# ══════════════════════════════════════════
# UI
# ══════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📦 Inventory Hub")
    st.markdown("---")
    week_label=st.text_input("Semana actual",value="WK13",label_visibility="collapsed")
    st.markdown(f"**Semana: {week_label}**")
    st.markdown("---")
    st.markdown("**Reglas Honduras:**")
    rules_hn=["Excluir Cut","Picked incluido","Is Second/Third → Irregulares","Locker Stock → VMI",
              "Sin VMI: Duluth, Tiltworks, Vortex","PO Open Order → Regular","Box Tag → Exceso/Obsoleto",
              "Create Date → Obsoleto/Exceso","FG vs Wip por Box Status"]
    for i,r in enumerate(rules_hn,1):
        st.markdown(f"<div style='font-size:.78rem;padding:2px 0;color:#a8c4e0;'><b style='color:#1B6CA8;'>{i}.</b> {r}</div>",unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("**Reglas TLP:**")
    rules_tlp=["Is Second/Third → TLP Irregulars","Blanks Excess → TLP Blanks Excess",
               "Printed Excess → TLP Printed Excess","Packed/Picked → TLP sin clasificacion","Resto → Wip"]
    for i,r in enumerate(rules_tlp,1):
        st.markdown(f"<div style='font-size:.78rem;padding:2px 0;color:#a8c4e0;'><b style='color:#1B6CA8;'>{i}.</b> {r}</div>",unsafe_allow_html=True)



# ══════════════════════════════════════════
# UI — Full App
# ══════════════════════════════════════════

import plotly.express as px
import plotly.graph_objects as go

import pandas as pd

def parse_prev_hn(df_raw):
    try:
        clas_map = {'Regulars':'Regular','VMI':'VMI','excess':'Exceso',
                    'Irregulars':'Irregulares','Obsolete':'Obsoleto','Liability':'Liability',
                    'Regular Wip':'Regular Wip','Wip':'Wip'}
        result = {}
        nrows, ncols = df_raw.shape

        # Step 1: Find the row containing week labels (WKnn)
        wk_row_idx = None
        val_col = None
        for ri in range(min(50, nrows)):
            row = df_raw.iloc[ri]
            for ci in range(ncols-1, -1, -1):
                cell = str(row.iloc[ci]).strip()
                if cell.upper().startswith('WK') and len(cell) > 2 and cell[2:].isdigit():
                    wk_row_idx = ri
                    val_col = ci
                    break
            if wk_row_idx is not None:
                break

        if val_col is None:
            return None

        lbl_col = val_col - 1

        # Step 2: Read data rows after the week label row
        for i in range(wk_row_idx + 1, min(wk_row_idx + 20, nrows)):
            row = df_raw.iloc[i]
            if lbl_col < 0 or val_col >= ncols: continue
            clas_raw = str(row.iloc[lbl_col]).strip() if pd.notna(row.iloc[lbl_col]) else ''
            val_raw  = str(row.iloc[val_col]).strip() if pd.notna(row.iloc[val_col]) else ''
            if clas_raw in clas_map and val_raw not in ['nan','-','','NaN']:
                try:
                    val = int(float(val_raw.replace(',','')))
                    result[clas_map[clas_raw]] = val
                except: pass

        return result if result else None
    except Exception as parse_err:
        import traceback
        print(f"parse_prev_hn ERROR: {parse_err}")
        print(traceback.format_exc())
        return None

def parse_prev_tlp(df_raw):
    try:
        clas_map = {'TLP irregulars':'Irregulares','TLP printed excess':'Exceso Printed',
                    'TLP sin clasificacion':'Sin Clasificacion','TLP Blanks excess':'Exceso Blanks'}
        result = {}
        headers = [str(h).strip() for h in df_raw.iloc[1].tolist()]
        for i in range(2, len(df_raw)):
            row = df_raw.iloc[i]
            client = str(row.iloc[0]).strip()
            if client in ['','nan','Grand Total','NaN']: continue
            for j, h in enumerate(headers[1:], 1):
                if h in clas_map and j < len(row):
                    val_raw = str(row.iloc[j]).strip()
                    if val_raw not in ['nan','-','','NaN']:
                        try:
                            val = int(float(val_raw.replace(',','')))
                            result[clas_map[h]] = result.get(clas_map[h], 0) + val
                        except: pass
        return result if result else None
    except:
        return None

def is_pivot_format(df):
    cols = [str(c).lower() for c in df.columns.tolist()]
    return 'clasificacion' not in cols and 'customer name' not in cols




import plotly.graph_objects as go
import re as _re

# ── Color constants ──
NAVY_UI = '#162447'
NAVY2 = '#1E3A6E'
LAVENDER = '#EEF2F7'
INDIGO = '#4F46E5'
INDIGO2 = '#6366F1'
SLATE = '#94A3B8'

CLAS_COLORS_HN = {
    'Regular': '#4F46E5', 'VMI': '#6366F1', 'Irregulares': '#F59E0B',
    'Exceso': '#F97316', 'Obsoleto': '#EF4444', 'Wip': '#94A3B8',
}
CLAS_COLORS_TLP = {
    'Sin Clasificacion': '#4F46E5', 'Irregulares': '#F59E0B',
    'Exceso Blanks': '#F97316', 'Exceso Printed': '#EF4444', 'Wip': '#94A3B8',
}

HN_FG_CLAS  = ['Regular','VMI','Irregulares','Exceso','Obsoleto']
HN_WIP_CLAS = ['Regular Wip']
TLP_FG_CLAS = ['Sin Clasificacion','Irregulares','Exceso Blanks','Exceso Printed']
TLP_WIP_CLAS= ['Wip']

def filter_df(df, view, fg_clas, wip_clas):
    if view == 'fg':
        if 'Type' in df.columns and df['Type'].notna().any():
            return df[df['Type']=='Finished Goods']
        return df[df['Clasificacion'].isin(fg_clas)]
    if view == 'wip':
        if 'Type' in df.columns and (df['Type']=='Wip').any():
            return df[df['Type']=='Wip']
        return df[df['Clasificacion'].isin(wip_clas)]
    if 'Type' in df.columns and df['Type'].notna().any():
        return df[df['Type'].isin(['Finished Goods','Wip'])]
    return df[df['Clasificacion'].isin(list(fg_clas) + list(wip_clas))]

def fmt(n):
    n = int(n)
    return f"{n:,}"

def fmtk(n):
    n = int(n)
    if n >= 1_000_000: return f"{n/1_000_000:.2f}M"
    if n >= 1_000: return f"{n/1_000:.0f}K"
    return str(n)

# ── Reusable HTML components ──
def kpi_card(label, value, sub='', top_color=INDIGO, sub_color='#818CF8'):
    return f"""<div style="background:#fff;border-radius:8px;padding:10px 12px;
border-top:3px solid {top_color};">
<div style="font-size:11px;color:#6366F1;text-transform:uppercase;letter-spacing:.07em;margin-bottom:6px;">{label}</div>
<div style="font-size:22px;font-weight:500;color:#162447;font-family:var(--font-mono);">{value}</div>
<div style="font-size:11px;color:{sub_color};margin-top:3px;">{sub}</div></div>"""

def clas_bar_row(name, val, total, color):
    pct = val/total*100 if total else 0
    bar_w = int(pct)
    return f"""<div style="display:flex;align-items:center;gap:6px;padding:5px 0;border-bottom:0.5px solid #EEF2F7;">
<div style="width:11px;height:11px;border-radius:50%;background:{color};flex-shrink:0;"></div>
<div style="flex:1;font-size:14px;color:#374151;">{name}</div>
<div style="width:80px;background:#EEF2F7;border-radius:2px;height:5px;">
  <div style="width:{bar_w}%;height:5px;border-radius:2px;background:{color};"></div>
</div>
<div style="font-size:14px;font-weight:500;color:#162447;min-width:44px;text-align:right;font-family:var(--font-mono);">{fmtk(val)}</div>
<div style="font-size:12px;color:#818CF8;min-width:36px;text-align:right;">{pct:.1f}%</div>
</div>"""

def top5_row(rank, code, val, max_val, color=INDIGO):
    bar_w = int(val/max_val*100) if max_val else 0
    return f"""<div style="display:flex;align-items:center;gap:6px;padding:4px 0;">
<div style="width:16px;height:16px;border-radius:50%;background:#EEF2F7;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:600;color:{color};flex-shrink:0;">{rank}</div>
<div style="font-size:13px;font-weight:600;color:#162447;min-width:30px;">{code}</div>
<div style="flex:1;background:#EEF2F7;border-radius:2px;height:5px;">
  <div style="width:{bar_w}%;height:5px;border-radius:2px;background:{color};"></div>
</div>
<div style="font-size:12px;color:#818CF8;min-width:42px;text-align:right;">{fmtk(val)}</div>
</div>"""

def alert_card_html(title, sub, icon, bg, rank_color, items, is_pct=False):
    rows = ""
    max_v = abs(items[0][1]) if items else 1
    for i,(code,val) in enumerate(items,1):
        bar_w = int(abs(val)/max_v*100) if max_v else 0
        sign = '+' if val >= 0 else ''
        val_str = f"{sign}{val:.1f}%" if is_pct else f"{sign}{int(val):,}"
        col = '#10B981' if val >= 0 else '#EF4444'
        rows += f"""<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:0.5px solid #EEF2F7;">
<div style="width:18px;height:18px;border-radius:50%;background:{bg};display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:600;color:{rank_color};flex-shrink:0;">{i}</div>
<div style="font-size:11px;font-weight:600;color:#162447;min-width:30px;">{code}</div>
<div style="flex:1;background:#EEF2F7;border-radius:2px;height:5px;"><div style="width:{bar_w}%;height:5px;border-radius:2px;background:{rank_color};opacity:.8;"></div></div>
<div style="font-size:10px;font-weight:500;color:{col};min-width:58px;text-align:right;">{val_str}</div></div>"""
    return f"""<div style="background:#fff;border-radius:8px;padding:10px 12px;height:100%;">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
<div style="width:26px;height:26px;border-radius:6px;background:{bg};display:flex;align-items:center;justify-content:center;font-size:12px;flex-shrink:0;">{icon}</div>
<div><div style="font-size:11px;font-weight:500;color:#162447;">{title}</div>
<div style="font-size:10px;color:#818CF8;">{sub}</div></div></div>{rows}</div>"""

def render_client_table(df, all_clas, color_map):
    df = df.copy()
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
    cajas = df.groupby('Customer Name').size()
    pivot = df.pivot_table(index='Customer Name', columns='Clasificacion', values='Quantity', aggfunc='sum', fill_value=0)
    for c in all_clas:
        if c not in pivot.columns: pivot[c] = 0
    pivot['_total'] = pivot[all_clas].sum(axis=1)
    pivot = pivot.sort_values('_total', ascending=False)
    total_uds = int(pivot['_total'].sum())

    th = f"padding:8px 10px;text-align:right;font-size:10px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.04em;background:#EEF2F7;"
    th_l = th.replace("text-align:right","text-align:left")
    td_s = "border-bottom:0.5px solid #EEF2F7;padding:7px 10px;font-size:11px;"

    headers = f'<th style="{th_l}">Código</th><th style="{th}">Cajas</th><th style="{th}">Total Uds</th><th style="{th}">% Total</th>'
    for c in all_clas:
        dot = color_map.get(c,'#888')
        headers += f'<th style="{th}"><span style="display:inline-block;width:7px;height:7px;border-radius:50%;background:{dot};margin-right:4px;vertical-align:middle;"></span>{c}</th>'

    rows_html = ""
    for i,(cli,row) in enumerate(pivot.iterrows()):
        bg = "background:#F8FAFF;" if i%2==0 else "background:#fff;"
        pct = row['_total']/total_uds*100 if total_uds else 0
        rows_html += f'<tr style="{bg}">'
        rows_html += f'<td style="{td_s}text-align:left;font-weight:600;">{cli}</td>'
        rows_html += f'<td style="{td_s}text-align:right;">{int(cajas.get(cli,0)):,}</td>'
        rows_html += f'<td style="{td_s}text-align:right;font-weight:600;">{int(row["_total"]):,}</td>'
        rows_html += f'<td style="{td_s}text-align:right;color:#818CF8;">{pct:.1f}%</td>'
        for c in all_clas:
            val = int(row.get(c,0))
            rows_html += f'<td style="{td_s}text-align:right;">{val:,}</td>' if val else f'<td style="{td_s}text-align:right;color:#C7D2FE;">—</td>'
        rows_html += '</tr>'

    tot_td = f"padding:8px 10px;font-size:11px;font-weight:600;color:{INDIGO};background:#EEF2F7;border-top:1px solid #C7D2FE;"
    rows_html += f'<tr><td style="{tot_td}text-align:left;">Total</td>'
    rows_html += f'<td style="{tot_td}text-align:right;">{int(cajas.sum()):,}</td>'
    rows_html += f'<td style="{tot_td}text-align:right;">{total_uds:,}</td>'
    rows_html += f'<td style="{tot_td}text-align:right;">100%</td>'
    for c in all_clas:
        t = int(pivot[c].sum())
        rows_html += f'<td style="{tot_td}text-align:right;">{t:,}</td>' if t else f'<td style="{tot_td}text-align:right;">—</td>'
    rows_html += '</tr>'

    st.markdown(
        f'<div style="overflow-x:auto;border-radius:8px;border:0.5px solid #C7D2FE;">'
        f'<table style="width:100%;border-collapse:collapse;font-family:var(--font-sans);white-space:nowrap;">'
        f'<thead><tr>{headers}</tr></thead><tbody>{rows_html}</tbody></table></div>',
        unsafe_allow_html=True)

def render_age_bars(df):
    df = df.copy()
    df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
    age = df.groupby('Year')['Quantity'].sum().sort_index()
    if age.empty: return
    max_v = age.max()
    colors = {y: ('#EF4444' if i==0 else '#F97316' if i==1 else '#F59E0B' if i==2 else '#6366F1' if i==3 else '#4F46E5' if i==4 else '#162447')
              for i,y in enumerate(sorted(age.index))}
    html = ""
    for yr, val in age.items():
        bar_w = int(val/max_v*100) if max_v else 0
        color = colors.get(yr, INDIGO)
        html += f"""<div style="display:flex;align-items:center;gap:8px;padding:4px 0;border-bottom:0.5px solid #EEF2F7;">
<div style="font-size:11px;color:#374151;min-width:36px;">{int(yr)}</div>
<div style="flex:1;background:#EEF2F7;border-radius:2px;height:7px;">
  <div style="width:{bar_w}%;height:7px;border-radius:2px;background:{color};"></div>
</div>
<div style="font-size:11px;font-weight:500;color:#162447;min-width:60px;text-align:right;font-family:var(--font-mono);">{int(val):,}</div></div>"""
    st.markdown(html, unsafe_allow_html=True)

def render_donut_plotly(data_series, color_map, height=420):
    labels = list(data_series.index)
    values = [int(v) for v in data_series.values]
    colors = [color_map.get(l,'#94A3B8') for l in labels]
    fig = go.Figure(go.Pie(
        labels=labels, values=values, marker_colors=colors,
        hole=0.50, textinfo='percent', textfont=dict(size=14),
        hovertemplate='%{label}<br>%{value:,}<extra></extra>'
    ))
    fig.update_layout(
        height=height, margin=dict(t=10,b=10,l=10,r=10),
        paper_bgcolor='rgba(0,0,0,0)',
        legend=dict(font=dict(size=13), orientation='v', x=1.0)
    )
    return fig

def render_analysis(r_cur, prev_df, prev_clas_dict, clas_colors, label):
    if r_cur is None:
        st.info(f"Clasifica {label} primero.")
        return

    r_cur = r_cur.copy()
    r_cur['Quantity'] = pd.to_numeric(r_cur['Quantity'], errors='coerce').fillna(0)
    cur_clas = r_cur.groupby('Clasificacion')['Quantity'].sum()

    prev_clas = pd.Series(dtype=float)
    has_client = False

    if prev_df is not None and 'Clasificacion' in prev_df.columns:
        prev_df = prev_df.copy()
        prev_df['Quantity'] = pd.to_numeric(prev_df['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
        prev_clas = prev_df.groupby('Clasificacion')['Quantity'].sum()
        has_client = 'Customer Name' in prev_df.columns
    elif prev_clas_dict:
        prev_clas = pd.Series(prev_clas_dict)
    else:
        st.warning("Sube el inventario de la semana anterior para ver el análisis.")
        return

    all_c = set(list(cur_clas.index)+list(prev_clas.index))
    clas_diff = {c: int(cur_clas.get(c,0)-prev_clas.get(c,0)) for c in all_c}
    top_clas = sorted(clas_diff.items(), key=lambda x: x[1], reverse=True)[:5]

    top_up = top_dn = top_pct = []
    if has_client:
        cur_cli  = r_cur.groupby('Customer Name')['Quantity'].sum()
        prev_cli = prev_df.groupby('Customer Name')['Quantity'].sum()
        all_cli  = set(list(cur_cli.index)+list(prev_cli.index))
        diffs    = {c: int(cur_cli.get(c,0)-prev_cli.get(c,0)) for c in all_cli}
        top_up   = sorted([(c,d) for c,d in diffs.items() if d>0], key=lambda x: x[1], reverse=True)[:5]
        top_dn   = sorted([(c,d) for c,d in diffs.items() if d<0], key=lambda x: x[1])[:5]
        top_pct  = sorted([(c,d/prev_cli.get(c,1)*100) for c,d in diffs.items() if prev_cli.get(c,0)>0],
                           key=lambda x: x[1], reverse=True)[:5]

    total_cur  = int(cur_clas.sum())
    total_prev = int(prev_clas.sum())
    total_diff = total_cur - total_prev
    sign = '+' if total_diff >= 0 else ''
    biggest_clas = max(clas_diff.items(), key=lambda x: x[1]) if clas_diff else None

    summary = f"El inventario de <b>{label}</b> varió <b>{sign}{total_diff:,} uds</b> esta semana ({fmtk(total_prev)} → {fmtk(total_cur)}). "
    if has_client and top_up:
        summary += f"El cliente <b>{top_up[0][0]}</b> fue el mayor contribuyente con <b>+{top_up[0][1]:,} uds</b>. "
    if has_client and top_dn:
        summary += f"El cliente <b>{top_dn[0][0]}</b> tuvo la mayor reducción con <b>{top_dn[0][1]:,} uds</b>. "
    if biggest_clas and biggest_clas[1] != 0:
        pct_c = biggest_clas[1]/prev_clas.get(biggest_clas[0],1)*100 if prev_clas.get(biggest_clas[0],0)>0 else 0
        summary += f"La clasificación <b>{biggest_clas[0]}</b> {'creció' if biggest_clas[1]>0 else 'bajó'} <b>{pct_c:+.0f}%</b>."
    if not has_client:
        summary += f" <span style='color:#F59E0B;font-size:10px;'>Sube el CSV de data completa para análisis por cliente.</span>"

    st.markdown(f"""<div style="background:#fff;border-radius:8px;padding:14px 16px;margin-bottom:16px;">
<div style="font-size:11px;font-weight:500;color:#162447;margin-bottom:8px;">Resumen automático</div>
<div style="font-size:12px;color:#374151;line-height:1.7;">{summary}</div></div>""", unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        if top_up:
            st.markdown(alert_card_html("Mayor incremento","Top 5 clientes que más subieron","▲","#DCFCE7","#166534",top_up), unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if top_pct:
            st.markdown(alert_card_html("Mayor % de crecimiento","Clientes que más crecieron proporcionalmente","!","#FEF9C3","#854D0E",top_pct,is_pct=True), unsafe_allow_html=True)
    with c2:
        if top_dn:
            st.markdown(alert_card_html("Mayor reducción","Top 5 clientes que más bajaron","▼","#FEE2E2","#991B1B",top_dn), unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if top_clas:
            rows = ""
            max_v = abs(top_clas[0][1]) if top_clas else 1
            for i,(clas,val) in enumerate(top_clas,1):
                bar_w = int(abs(val)/max_v*100) if max_v else 0
                color = clas_colors.get(clas,'#7C3AED')
                sign2 = '+' if val>=0 else ''
                rows += f"""<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:0.5px solid #EEF2F7;">
<div style="width:18px;height:18px;border-radius:50%;background:#EDE9FE;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:600;color:#5B21B6;flex-shrink:0;">{i}</div>
<div style="font-size:11px;font-weight:600;color:#162447;min-width:80px;">{clas}</div>
<div style="flex:1;background:#EEF2F7;border-radius:2px;height:5px;"><div style="width:{bar_w}%;height:5px;border-radius:2px;background:{color};opacity:.8;"></div></div>
<div style="font-size:10px;font-weight:500;color:#5B21B6;min-width:58px;text-align:right;">{sign2}{val:,}</div></div>"""
            st.markdown(f"""<div style="background:#fff;border-radius:8px;padding:10px 12px;">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
<div style="width:26px;height:26px;border-radius:6px;background:#EDE9FE;display:flex;align-items:center;justify-content:center;font-size:12px;">★</div>
<div><div style="font-size:11px;font-weight:500;color:#162447;">Clasificaciones críticas</div>
<div style="font-size:10px;color:#818CF8;">Las que más cambiaron esta semana</div></div></div>{rows}</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════
with st.sidebar:
    st.markdown("## Inventory Hub")
    st.caption("v3 — 2025-04-11")
    week_label = st.text_input("Semana", value="WK13", key="week_input")
    st.markdown(f"**Semana: {week_label}**")
    st.markdown("---")
    st.markdown("**Cargar archivos**")
    st.caption("La app detecta automáticamente cada archivo por su nombre.")
    all_files = st.file_uploader("Selecciona todos los CSVs", type=['csv'],
                                  accept_multiple_files=True, key='all_files')

    carton_file = open_file = prev_hn_file = tlp_file = prev_tlp_file = None
    for f in (all_files or []):
        n = f.name.lower()
        if 'carton' in n and 'tlp' in n: tlp_file = f
        elif 'carton' in n: carton_file = f
        elif 'open' in n and 'order' in n: open_file = f
        elif 'open_order' in n: open_file = f
        elif ('data_hn' in n or 'data_honduras' in n or 'fg_semana' in n
              or ('fg' in n and 'semana' in n)
              or ('semana' in n and 'tlp' not in n and 'wk' in n)): prev_hn_file = f
        elif ('data_tlp' in n or 'inventario_tlp' in n
              or ('tlp' in n and ('semana' in n or 'wk' in n))): prev_tlp_file = f

    if all_files:
        def chk(f, lbl):
            if f: st.markdown(f"<div style='font-size:11px;color:#A5B4FC;'>✓ {lbl}</div>", unsafe_allow_html=True)
            else: st.markdown(f"<div style='font-size:11px;color:#475569;'>— {lbl}: no detectado</div>", unsafe_allow_html=True)
        chk(carton_file,"Carton HN"); chk(open_file,"Open Order")
        chk(tlp_file,"Carton TLP"); chk(prev_hn_file,"HN anterior"); chk(prev_tlp_file,"TLP anterior")

    st.markdown("---")
    if st.button("Clasificar ambas bodegas", type="primary", use_container_width=True):
        if carton_file and open_file:
            with st.spinner("Clasificando Honduras..."):
                carton_file.seek(0); open_file.seek(0)
                r_hn, cut = classify_honduras(pd.read_csv(carton_file,low_memory=False), pd.read_csv(open_file,low_memory=False))
                st.session_state['hn_r'] = r_hn; st.session_state['hn_cut'] = cut
                if 'hist_hn' not in st.session_state: st.session_state['hist_hn'] = {}
                _wk = week_label.replace('WK','').replace('wk','').strip()
                st.session_state['hist_hn'][week_label] = r_hn.groupby('Clasificacion')['Quantity'].sum().to_dict()

        if prev_hn_file:
            prev_hn_file.seek(0)
            raw = pd.read_csv(prev_hn_file, low_memory=False, header=None)
            if is_pivot_format(raw):
                parsed = parse_prev_hn(raw)
                st.session_state['hn_prev_df'] = None
                st.session_state['hn_prev_clas'] = parsed or {}
                if 'hist_hn' not in st.session_state: st.session_state['hist_hn'] = {}
                _wkn = week_label.replace('WK','').replace('wk','').strip()
                prev_wk_h = f"WK{int(_wkn)-1}" if _wkn.isdigit() else "WK Ant."
                st.session_state['hist_hn'][prev_wk_h] = parsed if parsed else {}
                st.sidebar.caption(f"WK12 saved: {sum((parsed or {}).values()):,} uds ({len(parsed or {})} clas.)")
            else:
                prev_hn_file.seek(0)
                df_prev_h = pd.read_csv(prev_hn_file, low_memory=False)
                st.session_state['hn_prev_df'] = df_prev_h
                st.session_state['hn_prev_clas'] = None
                if 'Clasificacion' in df_prev_h.columns:
                    df_prev_h2 = df_prev_h.copy()
                    df_prev_h2['Quantity'] = pd.to_numeric(df_prev_h2['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    clas_h = df_prev_h2.groupby('Clasificacion')['Quantity'].sum().to_dict()
                    if 'hist_hn' not in st.session_state: st.session_state['hist_hn'] = {}
                    _wkn2 = week_label.replace('WK','').replace('wk','').strip()
                    prev_wk_h2 = f"WK{int(_wkn2)-1}" if _wkn2.isdigit() else "WK Ant."
                    st.session_state['hist_hn'][prev_wk_h2] = clas_h
        else:
            st.session_state['hn_prev_df'] = None
            st.session_state['hn_prev_clas'] = None

        if tlp_file:
            with st.spinner("Clasificando TLP..."):
                tlp_file.seek(0)
                r_tlp = classify_tlp(pd.read_csv(tlp_file,low_memory=False))
                st.session_state['tlp_r'] = r_tlp
                if 'hist_tlp' not in st.session_state: st.session_state['hist_tlp'] = {}
                st.session_state['hist_tlp'][week_label] = r_tlp.groupby('Clasificacion')['Quantity'].sum().to_dict()

        if prev_tlp_file:
            prev_tlp_file.seek(0)
            raw2 = pd.read_csv(prev_tlp_file, low_memory=False, header=None)
            if is_pivot_format(raw2):
                parsed2 = parse_prev_tlp(raw2)
                st.session_state['tlp_prev_df'] = None
                st.session_state['tlp_prev_clas'] = parsed2 or {}
                if 'hist_tlp' not in st.session_state: st.session_state['hist_tlp'] = {}
                _wknt = week_label.replace('WK','').replace('wk','').strip()
                prev_wk_t = f"WK{int(_wknt)-1}" if _wknt.isdigit() else "WK Ant."
                st.session_state['hist_tlp'][prev_wk_t] = parsed2 or {}
            else:
                prev_tlp_file.seek(0)
                df_prev_t = pd.read_csv(prev_tlp_file, low_memory=False)
                st.session_state['tlp_prev_df'] = df_prev_t
                st.session_state['tlp_prev_clas'] = None
                if 'Clasificacion' in df_prev_t.columns:
                    df_prev_t2 = df_prev_t.copy()
                    df_prev_t2['Quantity'] = pd.to_numeric(df_prev_t2['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    clas_t = df_prev_t2.groupby('Clasificacion')['Quantity'].sum().to_dict()
                    if 'hist_tlp' not in st.session_state: st.session_state['hist_tlp'] = {}
                    _wknt2 = week_label.replace('WK','').replace('wk','').strip()
                    prev_wk_t2 = f"WK{int(_wknt2)-1}" if _wknt2.isdigit() else "WK Ant."
                    st.session_state['hist_tlp'][prev_wk_t2] = clas_t
        else:
            st.session_state['tlp_prev_df'] = None
            st.session_state['tlp_prev_clas'] = None

        st.success("✓ Listo!")
        # Debug: show what got saved
        _h = st.session_state.get('hist_hn',{})
        st.caption(f"✓ hist_hn: {list(_h.keys())} | WK12={dict(list(_h.get('WK12',{}).items())[:2])}")

# ══ Session state ══
r_hn   = st.session_state.get('hn_r')
r_tlp  = st.session_state.get('tlp_r')
cut_n, cut_u = st.session_state.get('hn_cut',(0,0))
prev_hn  = st.session_state.get('hn_prev_df')
prev_tlp = st.session_state.get('tlp_prev_df')
prev_hn_clas  = st.session_state.get('hn_prev_clas') or {}
prev_tlp_clas = st.session_state.get('tlp_prev_clas') or {}
hist_hn  = st.session_state.get('hist_hn', {})
hist_tlp = st.session_state.get('hist_tlp', {})

def wk_sort(w):
    m = _re.search(r'(\d+)', str(w))
    return int(m.group(1)) if m else 999

# ══════════════════════════════════════════
# TABS
# ══════════════════════════════════════════
tab_dash, tab_hn, tab_tlp, tab_comp, tab_hist, tab_dl = st.tabs([
    "Dashboard", "Honduras", "TLP", "Comparativo", "Histórico", "Descargas"
])

# ══ TAB: DASHBOARD ══
with tab_dash:
    if r_hn is None and r_tlp is None:
        st.info("Carga los archivos en el panel izquierdo y presiona Clasificar.")
    else:
        vmap = {"Todo":"all","Finished Goods":"fg","Wip":"wip"}
        view_d = st.radio("", ["Todo","Finished Goods","Wip"], horizontal=True, key="dash_view", label_visibility="collapsed")
        df_vhn  = filter_df(r_hn,  vmap[view_d], HN_FG_CLAS, HN_WIP_CLAS) if r_hn  is not None else pd.DataFrame()
        df_vtlp = filter_df(r_tlp, vmap[view_d], TLP_FG_CLAS, TLP_WIP_CLAS) if r_tlp is not None else pd.DataFrame()
        if not df_vhn.empty:  df_vhn['Quantity']  = pd.to_numeric(df_vhn['Quantity'],  errors='coerce').fillna(0)
        if not df_vtlp.empty: df_vtlp['Quantity'] = pd.to_numeric(df_vtlp['Quantity'], errors='coerce').fillna(0)
        tot_hn  = int(df_vhn['Quantity'].sum())  if not df_vhn.empty  else 0
        tot_tlp = int(df_vtlp['Quantity'].sum()) if not df_vtlp.empty else 0
        tot_all = tot_hn + tot_tlp
        caj_hn  = len(df_vhn)  if not df_vhn.empty  else 0
        caj_tlp = len(df_vtlp) if not df_vtlp.empty else 0

        k1,k2,k3,k4,k5 = st.columns(5)
        with k1: st.markdown(kpi_card("Total ambas bodegas", fmtk(tot_all), f"{tot_all:,} uds", '#162447','#818CF8'), unsafe_allow_html=True)
        with k2: st.markdown(kpi_card("Honduras uds", fmtk(tot_hn), f"{caj_hn:,} cajas", INDIGO), unsafe_allow_html=True)
        with k3: st.markdown(kpi_card("Honduras %", f"{tot_hn/tot_all*100:.1f}%" if tot_all else "—", "del total", INDIGO2), unsafe_allow_html=True)
        with k4: st.markdown(kpi_card("TLP uds", fmtk(tot_tlp), f"{caj_tlp:,} cajas", SLATE,'#64748B'), unsafe_allow_html=True)
        with k5: st.markdown(kpi_card("TLP %", f"{tot_tlp/tot_all*100:.1f}%" if tot_all else "—", "del total", SLATE,'#64748B'), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Donuts
        col_hn, col_tlp = st.columns(2)
        with col_hn:
            if r_hn is not None and not df_vhn.empty:
                st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Honduras — clasificación</div>", unsafe_allow_html=True)
                cs = df_vhn.groupby('Clasificacion')['Quantity'].sum().sort_values(ascending=False)
                st.plotly_chart(render_donut_plotly(cs, CLAS_COLORS_HN), use_container_width=True)
                total_cs = int(cs.sum())
                html_det = ""
                for clas, val in cs.items():
                    html_det += clas_bar_row(clas, val, total_cs, CLAS_COLORS_HN.get(clas,'#94A3B8'))
                st.markdown(html_det, unsafe_allow_html=True)

        with col_tlp:
            if r_tlp is not None and not df_vtlp.empty:
                st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>TLP — clasificación</div>", unsafe_allow_html=True)
                cs2 = df_vtlp.groupby('Clasificacion')['Quantity'].sum().sort_values(ascending=False)
                st.plotly_chart(render_donut_plotly(cs2, CLAS_COLORS_TLP), use_container_width=True)
                total_cs2 = int(cs2.sum())
                html_det2 = ""
                for clas, val in cs2.items():
                    html_det2 += clas_bar_row(clas, val, total_cs2, CLAS_COLORS_TLP.get(clas,'#94A3B8'))
                st.markdown(html_det2, unsafe_allow_html=True)

        # Top 5 section
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:11px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Top 5 clientes — ambas bodegas</div>", unsafe_allow_html=True)
        t1, t2 = st.columns(2)
        with t1:
            if r_hn is not None and not df_vhn.empty:
                df_vhn['Quantity'] = pd.to_numeric(df_vhn['Quantity'], errors='coerce').fillna(0)
                top_hn = df_vhn.groupby('Customer Name')['Quantity'].sum().sort_values(ascending=False).head(5)
                max_v = int(top_hn.iloc[0]) if len(top_hn) else 1
                st.markdown(f"<div style='font-size:10px;color:#818CF8;margin-bottom:6px;'>Honduras</div>", unsafe_allow_html=True)
                html_t = "".join(top5_row(i+1, c, int(v), max_v, INDIGO) for i,(c,v) in enumerate(top_hn.items()))
                st.markdown(f"<div style='background:#fff;border-radius:8px;padding:10px 12px;'>{html_t}</div>", unsafe_allow_html=True)
        with t2:
            if r_tlp is not None and not df_vtlp.empty:
                df_vtlp['Quantity'] = pd.to_numeric(df_vtlp['Quantity'], errors='coerce').fillna(0)
                top_tlp = df_vtlp.groupby('Customer Name')['Quantity'].sum().sort_values(ascending=False).head(5)
                max_v2 = int(top_tlp.iloc[0]) if len(top_tlp) else 1
                st.markdown(f"<div style='font-size:10px;color:#818CF8;margin-bottom:6px;'>TLP</div>", unsafe_allow_html=True)
                html_t2 = "".join(top5_row(i+1, c, int(v), max_v2, SLATE) for i,(c,v) in enumerate(top_tlp.items()))
                st.markdown(f"<div style='background:#fff;border-radius:8px;padding:10px 12px;'>{html_t2}</div>", unsafe_allow_html=True)

# ══ TAB: HONDURAS ══
with tab_hn:
    if r_hn is None:
        st.info("Carga el Carton Report y Open Order en el panel izquierdo.")
    else:
        vmap2 = {"Todo":"all","Finished Goods":"fg","Wip":"wip"}
        view_h = st.radio("", ["Todo","Finished Goods","Wip"], horizontal=True, key="hn_view", label_visibility="collapsed")
        df_h = filter_df(r_hn, vmap2[view_h], HN_FG_CLAS, HN_WIP_CLAS).copy()
        df_h['Quantity'] = pd.to_numeric(df_h['Quantity'], errors='coerce').fillna(0)
        tot_h = int(df_h['Quantity'].sum()); caj_h = len(df_h)
        # Dynamic KPIs based on filter — show only what's in df_h (already filtered)
        r_hn_q = r_hn.copy(); r_hn_q['Quantity'] = pd.to_numeric(r_hn_q['Quantity'], errors='coerce').fillna(0)
        reg_h  = int(df_h[df_h['Clasificacion']=='Regular']['Quantity'].sum())
        vmi_h  = int(df_h[df_h['Clasificacion']=='VMI']['Quantity'].sum())
        irr_h  = int(df_h[df_h['Clasificacion']=='Irregulares']['Quantity'].sum())
        obs_h  = int(df_h[df_h['Clasificacion']=='Obsoleto']['Quantity'].sum())
        exc_h  = int(df_h[df_h['Clasificacion']=='Exceso']['Quantity'].sum())
        wip_h  = int(df_h[df_h['Clasificacion'].isin(HN_WIP_CLAS)]['Quantity'].sum())
        k1,k2,k3,k4,k5,k6 = st.columns(6)
        with k1: st.markdown(kpi_card("Total", fmtk(tot_h), f"{caj_h:,} cajas"), unsafe_allow_html=True)
        with k2: st.markdown(kpi_card("Regular", fmtk(reg_h), f"{reg_h/max(tot_h,1)*100:.1f}%", INDIGO, INDIGO), unsafe_allow_html=True)
        with k3: st.markdown(kpi_card("VMI", fmtk(vmi_h), f"{vmi_h/max(tot_h,1)*100:.1f}%", INDIGO2, INDIGO2), unsafe_allow_html=True)
        with k4: st.markdown(kpi_card("Irregulares", fmtk(irr_h), f"{irr_h/max(tot_h,1)*100:.1f}%",'#F59E0B','#F59E0B'), unsafe_allow_html=True)
        with k5: st.markdown(kpi_card("Obsoleto+Exceso", fmtk(obs_h+exc_h), f"{(obs_h+exc_h)/max(tot_h,1)*100:.1f}%",'#EF4444','#EF4444'), unsafe_allow_html=True)
        with k6: st.markdown(kpi_card("Wip", fmtk(wip_h), f"{wip_h/max(tot_h,1)*100:.1f}%", SLATE,'#64748B'), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Distribución por clasificación</div>", unsafe_allow_html=True)
            cs_h = df_h.groupby('Clasificacion')['Quantity'].sum().sort_values(ascending=False)
            total_cs_h = int(cs_h.sum())
            html_cs = "".join(clas_bar_row(n, v, total_cs_h, CLAS_COLORS_HN.get(n,'#94A3B8')) for n,v in cs_h.items())
            st.markdown(f"<div style='background:#fff;border-radius:8px;padding:12px;'>{html_cs}</div>", unsafe_allow_html=True)
        with c2:
            st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Antigüedad del inventario</div>", unsafe_allow_html=True)
            st.markdown("<div style='background:#fff;border-radius:8px;padding:12px;'>", unsafe_allow_html=True)
            render_age_bars(df_h)
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Por cliente — mayor a menor</div>", unsafe_allow_html=True)
        all_clas_h = HN_FG_CLAS if vmap2[view_h]=='fg' else HN_WIP_CLAS if vmap2[view_h]=='wip' else HN_FG_CLAS+HN_WIP_CLAS
        render_client_table(df_h, all_clas_h, CLAS_COLORS_HN)

# ══ TAB: TLP ══
with tab_tlp:
    if r_tlp is None:
        st.info("Carga el Carton Report TLP en el panel izquierdo.")
    else:
        view_t = st.radio("", ["Todo","Finished Goods","Wip"], horizontal=True, key="tlp_view", label_visibility="collapsed")
        df_t = filter_df(r_tlp, vmap2[view_t], TLP_FG_CLAS, TLP_WIP_CLAS).copy()
        df_t['Quantity'] = pd.to_numeric(df_t['Quantity'], errors='coerce').fillna(0)
        df_t['Quantity'] = pd.to_numeric(df_t['Quantity'], errors='coerce').fillna(0)
        tot_t = int(df_t['Quantity'].sum()); caj_t = len(df_t)
        # Dynamic KPIs — calculated from filtered df_t
        sinc_t = int(df_t[df_t['Clasificacion']=='Sin Clasificacion']['Quantity'].sum())
        irr_t  = int(df_t[df_t['Clasificacion']=='Irregulares']['Quantity'].sum())
        exb_t  = int(df_t[df_t['Clasificacion']=='Exceso Blanks']['Quantity'].sum())
        exp_t  = int(df_t[df_t['Clasificacion']=='Exceso Printed']['Quantity'].sum())
        wip_t  = int(df_t[df_t['Clasificacion'].isin(TLP_WIP_CLAS)]['Quantity'].sum())
        k1,k2,k3,k4,k5,k6 = st.columns(6)
        with k1: st.markdown(kpi_card("Total", fmtk(tot_t), f"{caj_t:,} cajas"), unsafe_allow_html=True)
        with k2: st.markdown(kpi_card("Sin Clasificacion", fmtk(sinc_t), f"{sinc_t/max(tot_t,1)*100:.1f}%", INDIGO, INDIGO), unsafe_allow_html=True)
        with k3: st.markdown(kpi_card("Irregulares", fmtk(irr_t), f"{irr_t/max(tot_t,1)*100:.1f}%",'#F59E0B','#F59E0B'), unsafe_allow_html=True)
        with k4: st.markdown(kpi_card("Exceso Blanks", fmtk(exb_t), f"{exb_t/max(tot_t,1)*100:.1f}%",'#F97316','#F97316'), unsafe_allow_html=True)
        with k5: st.markdown(kpi_card("Exceso Printed", fmtk(exp_t), f"{exp_t/max(tot_t,1)*100:.1f}%",'#EF4444','#EF4444'), unsafe_allow_html=True)
        with k6: st.markdown(kpi_card("Wip", fmtk(wip_t), f"{wip_t/max(tot_t,1)*100:.1f}%", SLATE,'#64748B'), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Distribución por clasificación</div>", unsafe_allow_html=True)
            cs_t = df_t.groupby('Clasificacion')['Quantity'].sum().sort_values(ascending=False)
            total_cs_t = int(cs_t.sum())
            html_cst = "".join(clas_bar_row(n, v, total_cs_t, CLAS_COLORS_TLP.get(n,'#94A3B8')) for n,v in cs_t.items())
            st.markdown(f"<div style='background:#fff;border-radius:8px;padding:12px;'>{html_cst}</div>", unsafe_allow_html=True)
        with c2:
            st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Antigüedad del inventario</div>", unsafe_allow_html=True)
            st.markdown("<div style='background:#fff;border-radius:8px;padding:12px;'>", unsafe_allow_html=True)
            render_age_bars(df_t)
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Por cliente — mayor a menor</div>", unsafe_allow_html=True)
        all_clas_t = TLP_FG_CLAS if vmap2[view_t]=='fg' else TLP_WIP_CLAS if vmap2[view_t]=='wip' else TLP_FG_CLAS+TLP_WIP_CLAS
        render_client_table(df_t, all_clas_t, CLAS_COLORS_TLP)

# ══ TAB: COMPARATIVO (4 semanas) ══
with tab_comp:
    def render_comp_4wk(bodega_key, r_cur, hist, clas_colors, line_color, label, week_lbl):
        view_c = st.radio("Ver:", ["Todo","Finished Goods","Wip"], horizontal=True,
                          key=f"comp_{bodega_key}_view", label_visibility="visible")

        # Get last 4 weeks from hist
        all_wks = sorted(hist.keys(), key=wk_sort)
        last4   = all_wks[-4:] if len(all_wks) >= 4 else all_wks

        if not last4:
            st.info(f"Clasifica {label} y sube semanas anteriores para ver el comparativo.")
            return

        def filter_hist(d, view, fg_clas, wip_clas):
            if not d: return {}
            if view == 'fg':  return {k:v for k,v in d.items() if k in fg_clas}
            if view == 'wip': return {k:v for k,v in d.items() if k in wip_clas}
            return d

        fg_c   = HN_FG_CLAS  if bodega_key=='hn' else TLP_FG_CLAS
        wip_c  = HN_WIP_CLAS if bodega_key=='hn' else TLP_WIP_CLAS
        vmap_c = {"Todo":"all","Finished Goods":"fg","Wip":"wip"}
        weeks_data = [filter_hist(hist.get(w,{}), vmap_c.get(view_c,"all"), fg_c, wip_c) for w in last4]
        totals     = [sum(d.values()) for d in weeks_data]


        # Badges
        html_b = '<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">'
        for i,(w,t) in enumerate(zip(last4,totals)):
            is_cur = (w == week_lbl)
            border = f"border:2px solid {INDIGO};" if is_cur else "border:0.5px solid #C7D2FE;"
            if i == 0:
                diff_html = f'<div style="font-size:10px;color:#818CF8;">inicio</div>'
            else:
                diff = t - totals[i-1]
                pct  = diff/totals[i-1]*100 if totals[i-1] else 0
                col  = '#10B981' if diff >= 0 else '#EF4444'
                arr  = '▲' if diff >= 0 else '▼'
                sign = '+' if diff >= 0 else ''
                diff_html = f'<div style="font-size:10px;color:{col};font-weight:500;">{arr} {sign}{diff:,}<br><span style="font-size:9px;">{sign}{pct:.1f}%</span></div>'
            lbl_str = f"{w} — actual" if is_cur else w
            html_b += f'<div style="background:#fff;{border}border-radius:10px;padding:8px 14px;text-align:center;min-width:80px;flex:1;">'
            html_b += f'<div style="font-size:10px;color:#818CF8;font-weight:500;margin-bottom:3px;">{lbl_str}</div>'
            html_b += f'<div style="font-size:14px;font-weight:500;color:#162447;font-family:var(--font-mono);">{fmtk(t)}</div>'
            html_b += diff_html + '</div>'
        html_b += '</div>'
        st.markdown(html_b, unsafe_allow_html=True)

        # KPI row
        if len(totals) > 1:
            diff_last = totals[-1]-totals[-2]
            diff_total= totals[-1]-totals[0]
            dc1 = '#10B981' if diff_last>=0 else '#EF4444'
            dc2 = '#10B981' if diff_total>=0 else '#EF4444'
            k1,k2,k3 = st.columns(3)
            with k1: st.markdown(kpi_card("Semana actual", fmtk(totals[-1]), week_lbl, line_color, line_color), unsafe_allow_html=True)
            with k2: st.markdown(kpi_card(f"Var. {last4[-2]}→{last4[-1]}", f"{'+' if diff_last>=0 else ''}{diff_last:,}", f"{diff_last/totals[-2]*100:+.1f}%" if totals[-2] else "", dc1, dc1), unsafe_allow_html=True)
            with k3: st.markdown(kpi_card(f"Var. {last4[0]}→{last4[-1]}", f"{'+' if diff_total>=0 else ''}{diff_total:,}", "acumulado", dc2, dc2), unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

        # Line chart
        if len(last4) > 1:
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=last4, y=totals, mode='lines+markers',
                line=dict(color=line_color, width=2.5),
                marker=dict(size=8, color=line_color),
                fill='tozeroy',
                fillcolor=f"rgba({int(line_color[1:3],16)},{int(line_color[3:5],16)},{int(line_color[5:7],16)},0.08)",
                hovertemplate='%{x}: %{y:,} uds<extra></extra>'
            ))
            fig.update_layout(
                height=180, margin=dict(t=10,b=10,l=10,r=10),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(tickfont=dict(size=11), showgrid=False),
                yaxis=dict(tickfont=dict(size=10), tickformat=',', gridcolor='rgba(0,0,0,0.05)'),
                showlegend=False
            )
            st.plotly_chart(fig, use_container_width=True)

        # Table
        st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Tabla comparativa</div>", unsafe_allow_html=True)
        all_c = sorted(set(k for d in weeks_data for k in d.keys()))
        rows = []
        for c in all_c:
            row = {'Clasificación': c}
            for i,(w,d) in enumerate(zip(last4,weeks_data)):
                val = int(d.get(c,0))
                if i > 0:
                    prev_val = int(weeks_data[i-1].get(c,0))
                    arr = ' ▲' if val > prev_val else ' ▼' if val < prev_val else ''
                    col = '#10B981' if val > prev_val else '#EF4444' if val < prev_val else '#374151'
                    row[w] = f"{val:,}{arr}"
                else:
                    row[w] = f"{val:,}" if val else "—"
            if len(weeks_data) > 1:
                diff_c = int(weeks_data[-1].get(c,0)) - int(weeks_data[0].get(c,0))
                row['Var. total'] = f"{'+' if diff_c>=0 else ''}{diff_c:,}"
            rows.append(row)

        if rows:
            tot_row = {'Clasificación': 'Total'}
            for w,d in zip(last4,weeks_data):
                tot_row[w] = f"{int(sum(d.values())):,}"
            if len(weeks_data) > 1:
                tot_diff = int(sum(weeks_data[-1].values())) - int(sum(weeks_data[0].values()))
                tot_row['Var. total'] = f"{'+' if tot_diff>=0 else ''}{tot_diff:,}"
            rows.append(tot_row)
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # Alerts
        if len(last4) > 1:
            st.markdown("---")
            st.markdown(f"<div style='font-size:11px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Alertas — {last4[-2]} → {last4[-1]}</div>", unsafe_allow_html=True)
            prev_df_a = st.session_state.get(f'{bodega_key}_prev_df')
            prev_clas_a = st.session_state.get(f'{bodega_key}_prev_clas') or {}
            render_analysis(r_cur, prev_df_a, prev_clas_a, clas_colors, label)

    comp_hn, comp_tlp, comp_both = st.tabs(["Honduras","TLP","Ambas Bodegas"])

    with comp_hn:
        if r_hn is None: st.info("Clasifica Honduras primero.")
        else:
            _dbg_hist = st.session_state.get('hist_hn',{})
            st.caption(f"Debug hist_hn keys: {list(_dbg_hist.keys())} | WK12 data: {dict(list(_dbg_hist.get('WK12',{}).items())[:3])}")
            render_comp_4wk('hn', r_hn, _dbg_hist, CLAS_COLORS_HN, INDIGO, 'Honduras', week_label)

    with comp_tlp:
        if r_tlp is None: st.info("Clasifica TLP primero.")
        else: render_comp_4wk('tlp', r_tlp, st.session_state.get('hist_tlp',{}), CLAS_COLORS_TLP, SLATE, 'TLP', week_label)

    with comp_both:
        h_hn  = st.session_state.get('hist_hn',{})
        h_tlp = st.session_state.get('hist_tlp',{})
        all_wks_b = sorted(set(list(h_hn.keys())+list(h_tlp.keys())), key=wk_sort)
        if not all_wks_b:
            st.info("Clasifica ambas bodegas para ver el comparativo.")
        else:
            view_b = st.radio("Ver:", ["Todo","Finished Goods","Wip"], horizontal=True, key="comp_both_view")
            vmap_b = {"Todo":"all","Finished Goods":"fg","Wip":"wip"}
            vb = vmap_b.get(view_b,"all")

            def filt_d(d, v, fg, wip):
                if not d: return {}
                if v=='fg':  return {k:val for k,val in d.items() if k in fg}
                if v=='wip': return {k:val for k,val in d.items() if k in wip}
                return d

            hn_tots  = [sum(filt_d(h_hn.get(w,{}),  vb, HN_FG_CLAS,  HN_WIP_CLAS).values())  for w in all_wks_b]
            tlp_tots = [sum(filt_d(h_tlp.get(w,{}), vb, TLP_FG_CLAS, TLP_WIP_CLAS).values()) for w in all_wks_b]

            k1,k2,k3,k4 = st.columns(4)
            with k1: st.markdown(kpi_card("Honduras actual", fmtk(hn_tots[-1]) if hn_tots else "—", week_label, INDIGO), unsafe_allow_html=True)
            with k2: st.markdown(kpi_card("TLP actual", fmtk(tlp_tots[-1]) if tlp_tots else "—", week_label, SLATE,'#64748B'), unsafe_allow_html=True)
            if len(hn_tots)>1:
                dh = hn_tots[-1]-hn_tots[-2]; dt = tlp_tots[-1]-tlp_tots[-2]
                with k3: st.markdown(kpi_card("Var. HN sem.", f"{'+' if dh>=0 else ''}{dh:,}", f"{dh/hn_tots[-2]*100:+.1f}%" if hn_tots[-2] else "", '#10B981' if dh>=0 else '#EF4444','#10B981' if dh>=0 else '#EF4444'), unsafe_allow_html=True)
                with k4: st.markdown(kpi_card("Var. TLP sem.", f"{'+' if dt>=0 else ''}{dt:,}", f"{dt/tlp_tots[-2]*100:+.1f}%" if tlp_tots[-2] else "", '#10B981' if dt>=0 else '#EF4444','#10B981' if dt>=0 else '#EF4444'), unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            fig_b = go.Figure()
            if any(v>0 for v in hn_tots):
                fig_b.add_trace(go.Scatter(x=all_wks_b, y=hn_tots, name='Honduras', mode='lines+markers',
                    line=dict(color=INDIGO, width=2.5), marker=dict(size=7), hovertemplate='%{x}: %{y:,}<extra>Honduras</extra>'))
            if any(v>0 for v in tlp_tots):
                fig_b.add_trace(go.Scatter(x=all_wks_b, y=tlp_tots, name='TLP', mode='lines+markers',
                    line=dict(color=SLATE, width=2.5), marker=dict(size=7), hovertemplate='%{x}: %{y:,}<extra>TLP</extra>'))
            fig_b.update_layout(height=260, margin=dict(t=10,b=10,l=10,r=10),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                legend=dict(font=dict(size=11), orientation='h', y=1.1),
                xaxis=dict(tickfont=dict(size=11), showgrid=False),
                yaxis=dict(tickfont=dict(size=10), tickformat=',', gridcolor='rgba(0,0,0,0.05)'))
            st.plotly_chart(fig_b, use_container_width=True)

            rows_b = []
            for i,w in enumerate(all_wks_b):
                hn_t  = int(sum(filt_d(h_hn.get(w,{}),  vb, HN_FG_CLAS,  HN_WIP_CLAS).values()))
                tlp_t = int(sum(filt_d(h_tlp.get(w,{}), vb, TLP_FG_CLAS, TLP_WIP_CLAS).values()))
                row = {'Semana':w, 'Honduras':f"{hn_t:,}", 'TLP':f"{tlp_t:,}", 'Total':f"{hn_t+tlp_t:,}"}
                if i > 0:
                    prev_hn  = int(sum(filt_d(h_hn.get(all_wks_b[i-1],{}),  vb, HN_FG_CLAS,  HN_WIP_CLAS).values()))
                    prev_tlp = int(sum(filt_d(h_tlp.get(all_wks_b[i-1],{}), vb, TLP_FG_CLAS, TLP_WIP_CLAS).values()))
                    row['▲▼ HN']  = f"{'+' if hn_t-prev_hn>=0 else ''}{hn_t-prev_hn:,}"
                    row['▲▼ TLP'] = f"{'+' if tlp_t-prev_tlp>=0 else ''}{tlp_t-prev_tlp:,}"
                    row['▲▼ Tot'] = f"{'+' if (hn_t+tlp_t)-(prev_hn+prev_tlp)>=0 else ''}{(hn_t+tlp_t)-(prev_hn+prev_tlp):,}"
                else:
                    row['▲▼ HN'] = '—'; row['▲▼ TLP'] = '—'; row['▲▼ Tot'] = '—'
                rows_b.append(row)
            st.dataframe(pd.DataFrame(rows_b), use_container_width=True, hide_index=True)

# ══ TAB: HISTÓRICO ══
with tab_hist:
    hist_extra = st.file_uploader("Sube semanas adicionales (CSV data completa o resumen pivot)",
                                   type=['csv'], accept_multiple_files=True, key='hist_extra')
    if hist_extra:
        for hf in hist_extra:
            hname = hf.name.lower()
            m = _re.search(r'wk(\d+)', hname) or _re.search(r'(\d+)', hname)
            wk_lbl2 = f"WK{m.group(1)}" if m else hf.name[:8]
            hf.seek(0)
            raw_h = pd.read_csv(hf, low_memory=False, header=None)
            is_tlp_h = 'tlp' in hname
            if is_pivot_format(raw_h):
                p = parse_prev_tlp(raw_h) if is_tlp_h else parse_prev_hn(raw_h)
                if p:
                    hk = 'hist_tlp' if is_tlp_h else 'hist_hn'
                    if hk not in st.session_state: st.session_state[hk] = {}
                    st.session_state[hk][wk_lbl2] = p
            else:
                hf.seek(0)
                df_hh = pd.read_csv(hf, low_memory=False)
                if 'Clasificacion' in df_hh.columns and 'Quantity' in df_hh.columns:
                    df_hh['Quantity'] = pd.to_numeric(df_hh['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    ct = df_hh.groupby('Clasificacion')['Quantity'].sum().to_dict()
                    hk = 'hist_tlp' if is_tlp_h else 'hist_hn'
                    if hk not in st.session_state: st.session_state[hk] = {}
                    st.session_state[hk][wk_lbl2] = ct

    hist_hn2  = st.session_state.get('hist_hn', {})
    hist_tlp2 = st.session_state.get('hist_tlp', {})

    hist_tabs = st.tabs(["Honduras","TLP","Ambas Bodegas"])

    def render_hist_tab(hist, clas_colors, all_clas, line_color, label):
        if not hist:
            st.info(f"No hay datos históricos para {label}. Clasifica y sube semanas anteriores.")
            return

        wks_h = sorted(hist.keys(), key=wk_sort)
        tots_h = [sum(hist[w].values()) for w in wks_h]
        max_t = max(tots_h) if tots_h else 1
        min_t = min(tots_h) if tots_h else 0
        avg_t = int(sum(tots_h)/len(tots_h)) if tots_h else 0
        wk_max = wks_h[tots_h.index(max_t)] if tots_h else "—"
        wk_min = wks_h[tots_h.index(min_t)] if tots_h else "—"

        k1,k2,k3,k4 = st.columns(4)
        with k1: st.markdown(kpi_card("Semanas cargadas", str(len(wks_h)), f"{wks_h[0]} → {wks_h[-1]}" if wks_h else ""), unsafe_allow_html=True)
        with k2: st.markdown(kpi_card("Semana más alta", wk_max, fmtk(max_t),'#10B981','#10B981'), unsafe_allow_html=True)
        with k3: st.markdown(kpi_card("Semana más baja", wk_min, fmtk(min_t),'#EF4444','#EF4444'), unsafe_allow_html=True)
        with k4: st.markdown(kpi_card("Promedio semanal", fmtk(avg_t), "por semana",'#F59E0B','#F59E0B'), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Filter using Streamlit multiselect - actually interactive
        selected = st.multiselect(
            "Ver clasificaciones:",
            options=['Total'] + all_clas,
            default=['Total'] + all_clas,
            key=f"hist_{label}_filter"
        )
        if not selected: selected = ['Total'] + all_clas

        # Chart - only show selected
        fig_h = go.Figure()
        if 'Total' in selected:
            fig_h.add_trace(go.Scatter(x=wks_h, y=tots_h, name='Total', mode='lines+markers',
                line=dict(color='#162447', width=2.5, dash='dot'), marker=dict(size=4),
                hovertemplate='%{x}: %{y:,}<extra>Total</extra>'))
        for c in all_clas:
            if c not in selected: continue
            vals = [int(hist[w].get(c,0)) for w in wks_h]
            color = clas_colors.get(c,'#94A3B8')
            fig_h.add_trace(go.Scatter(x=wks_h, y=vals, name=c, mode='lines+markers',
                line=dict(color=color, width=1.5), marker=dict(size=3),
                hovertemplate=f'%{{x}}: %{{y:,}}<extra>{c}</extra>'))
        fig_h.update_layout(
            height=240, margin=dict(t=10,b=10,l=10,r=10),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            legend=dict(font=dict(size=10), orientation='h', y=1.1),
            xaxis=dict(tickfont=dict(size=10), showgrid=False),
            yaxis=dict(tickfont=dict(size=10), tickformat=',', gridcolor='rgba(0,0,0,0.05)'))
        st.plotly_chart(fig_h, use_container_width=True)

        # Summary table
        st.markdown(f"<div style='font-size:13px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:10px;'>Resumen por semana</div>", unsafe_allow_html=True)
        rows_h = []
        for i,w in enumerate(wks_h):
            d = hist[w]
            row = {'Semana': w}
            for c in all_clas:
                row[c] = int(d.get(c,0))
            row['Total'] = int(sum(d.values()))
            if i > 0:
                prev_tot = int(sum(hist[wks_h[i-1]].values()))
                diff_t = row['Total'] - prev_tot
                row['▲▼'] = f"{'+' if diff_t>=0 else ''}{diff_t:,}"
            else:
                row['▲▼'] = '—'
            rows_h.append(row)
        st.dataframe(pd.DataFrame(rows_h), use_container_width=True, hide_index=True)

    with hist_tabs[0]:
        render_hist_tab(hist_hn2, CLAS_COLORS_HN, HN_FG_CLAS+HN_WIP_CLAS, INDIGO, 'Honduras')
    with hist_tabs[1]:
        render_hist_tab(hist_tlp2, CLAS_COLORS_TLP, TLP_FG_CLAS+TLP_WIP_CLAS, SLATE, 'TLP')
    with hist_tabs[2]:
        all_wks_h = sorted(set(list(hist_hn2.keys())+list(hist_tlp2.keys())), key=wk_sort)
        if not all_wks_h:
            st.info("No hay datos históricos todavía.")
        else:
            hn_h  = [sum(hist_hn2.get(w,{}).values())  for w in all_wks_h]
            tlp_h = [sum(hist_tlp2.get(w,{}).values()) for w in all_wks_h]
            fig_bh = go.Figure()
            fig_bh.add_trace(go.Scatter(x=all_wks_h, y=hn_h,  name='Honduras', mode='lines+markers',
                line=dict(color=INDIGO, width=2), marker=dict(size=5)))
            fig_bh.add_trace(go.Scatter(x=all_wks_h, y=tlp_h, name='TLP', mode='lines+markers',
                line=dict(color=SLATE, width=2), marker=dict(size=5)))
            fig_bh.update_layout(height=260, margin=dict(t=10,b=10,l=10,r=10),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                legend=dict(font=dict(size=11), orientation='h', y=1.1),
                xaxis=dict(tickfont=dict(size=10), showgrid=False),
                yaxis=dict(tickfont=dict(size=10), tickformat=',', gridcolor='rgba(0,0,0,0.05)'))
            st.plotly_chart(fig_bh, use_container_width=True)

# ══ TAB: DESCARGAS ══
with tab_dl:
    st.markdown(f"<div style='font-size:11px;font-weight:500;color:{INDIGO};text-transform:uppercase;letter-spacing:.07em;margin-bottom:12px;'>Descargar archivos</div>", unsafe_allow_html=True)

    if r_hn is None and r_tlp is None:
        st.info("Clasifica primero los inventarios para poder descargar.")
    else:
      try:
        c1, c2 = st.columns(2)
        with c1:
            if r_hn is not None:
                st.markdown(f"<div style='font-size:11px;font-weight:500;color:{INDIGO};margin-bottom:8px;'>Honduras — {week_label}</div>", unsafe_allow_html=True)
                wk_prev_hn = None
                if prev_hn is not None and hasattr(prev_hn, 'columns') and 'Clasificacion' in prev_hn.columns and 'Customer Name' in prev_hn.columns:
                    prev_hn4 = prev_hn.copy()
                    prev_hn4['Quantity'] = pd.to_numeric(prev_hn4['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    inv_types = ['Regulars','VMI','Excess','Irregulars','Obsolete','Liability']
                    prev_hn4['Clas_Col'] = prev_hn4['Clasificacion'].map(CMAP_HN)
                    pfg = prev_hn4[prev_hn4['Type']=='Finished Goods'] if 'Type' in prev_hn4.columns else prev_hn4
                    pp = pfg.pivot_table(index='Customer Name',columns='Clas_Col',values='Quantity',aggfunc='sum',fill_value=0)
                    wk_prev_hn = {cli:{t:int(pp.loc[cli,t]) if cli in pp.index and t in pp.columns else 0 for t in inv_types} for cli in ACTIVOS+INACTIVOS}
                try:
                    with st.spinner("Generando Excel..."):
                        buf_hn = build_excel_hn(r_hn, wk_prev_hn, week_label)
                    st.download_button(f"Excel Honduras {week_label}", data=buf_hn,
                    file_name=f"inventario_honduras_{week_label}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                    b2 = io.StringIO(); r_hn.to_csv(b2, index=False)
                    st.download_button("CSV Honduras completo", data=b2.getvalue(),
                        file_name=f"data_honduras_{week_label}.csv", mime="text/csv",
                        use_container_width=True)
                except Exception as e:
                    st.error(f"Error generando Excel Honduras: {str(e)}")
                    b2 = io.StringIO(); r_hn.to_csv(b2, index=False)
                    st.download_button("CSV Honduras completo", data=b2.getvalue(),
                        file_name=f"data_honduras_{week_label}.csv", mime="text/csv",
                        use_container_width=True)

        with c2:
            if r_tlp is not None:
                st.markdown(f"<div style='font-size:11px;font-weight:500;color:{INDIGO};margin-bottom:8px;'>TLP — {week_label}</div>", unsafe_allow_html=True)
                wk_prev_tlp = None
                if prev_tlp is not None and hasattr(prev_tlp,'columns') and 'Clasificacion' in prev_tlp.columns and 'Customer Name' in prev_tlp.columns:
                    prev_tlp4 = prev_tlp.copy()
                    prev_tlp4['Quantity'] = pd.to_numeric(prev_tlp4['Quantity'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
                    inv_comp = ['TLP Irregulars','TLP Printed Excess','TLP sin clasificacion','TLP Blanks Excess']
                    pp2 = prev_tlp4.pivot_table(index='Customer Name',columns='Clasificacion',values='Quantity',aggfunc='sum',fill_value=0)
                    all_tlp_c = list(set(TLP_ORDER+list(pp2.index)))
                    wk_prev_tlp = {cli:{t:int(pp2.loc[cli,t]) if cli in pp2.index and t in pp2.columns else 0 for t in inv_comp} for cli in all_tlp_c}
                try:
                    with st.spinner("Generando Excel..."):
                        buf_tlp = build_excel_tlp(r_tlp, wk_prev_tlp, week_label)
                    st.download_button(f"Excel TLP {week_label}", data=buf_tlp,
                        file_name=f"inventario_TLP_{week_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.error(f"Error generando Excel TLP: {str(e)}")
                b4 = io.StringIO(); r_tlp.to_csv(b4, index=False)
                st.download_button("CSV TLP completo", data=b4.getvalue(),
                    file_name=f"data_TLP_{week_label}.csv", mime="text/csv",
                    use_container_width=True)
      except Exception as e_dl:
          st.error(f'Error en Descargas: {e_dl}')
